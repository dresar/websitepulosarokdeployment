import csv
import io
import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.core.paginator import Paginator
from django.db.models import Q
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# from references.models import Penduduk, Dusun, Lorong, RT, RW, Keluarga  # COMMENTED OUT - references app disabled
# Using letters app models instead
try:
    from letters.models import Penduduk, Dusun, Lorong, RT, RW, Keluarga
except ImportError:
    Penduduk = None
    Dusun = None
    Lorong = None
    RT = None
    RW = None
    Keluarga = None
from .template_generators import get_import_template


class ExportImportMixin:
    """Mixin untuk export/import functionality"""
    
    def get_export_data(self, queryset, format_type='excel'):
        """Generate export data berdasarkan format"""
        if format_type == 'csv':
            return self.export_csv(queryset)
        elif format_type == 'excel':
            return self.export_excel(queryset)
        elif format_type == 'pdf':
            return self.export_pdf(queryset)
        elif format_type == 'template':
            return self.export_template()
        else:
            return self.export_excel(queryset)  # Default ke Excel
    
    def export_csv(self, queryset):
        """Export data ke CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="penduduk_data.csv"'
        
        # BOM untuk UTF-8 (agar Excel bisa baca dengan benar)
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Header
        writer.writerow([
            'NIK', 'Nama Lengkap', 'Jenis Kelamin', 'Tempat Lahir', 'Tanggal Lahir',
            'Agama', 'Pendidikan Terakhir', 'Pekerjaan', 'Status Perkawinan',
            'Golongan Darah', 'Dusun', 'Lorong', 'Alamat', 'RT', 'RW',
            'No. Telepon', 'No. Handphone', 'Email', 'No. KK'
        ])
        
        # Data
        for penduduk in queryset:
            writer.writerow([
                penduduk.nik,
                penduduk.name,
                penduduk.get_gender_display(),
                penduduk.birth_place or '',
                penduduk.birth_date.strftime('%d/%m/%Y') if penduduk.birth_date else '',
                penduduk.religion or '',
                penduduk.get_education_display() if penduduk.education else '',
                penduduk.occupation or '',
                penduduk.get_marital_status_display() if penduduk.marital_status else '',
                penduduk.blood_type or '',
                penduduk.dusun.name if penduduk.dusun else '',
                penduduk.lorong.nama_lorong if penduduk.lorong else '',
                penduduk.address or '',
                penduduk.rt_number or '',
                penduduk.rw_number or '',
                penduduk.phone_number or '',
                penduduk.mobile_number or '',
                penduduk.email or '',
                penduduk.kk_number or ''
            ])
        
        return response
    
    def export_excel(self, queryset):
        """Export data ke Excel dengan styling"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="penduduk_data.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Penduduk"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border styling
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'No', 'NIK', 'Nama Lengkap', 'Jenis Kelamin', 'Tempat Lahir', 'Tanggal Lahir',
            'Agama', 'Pendidikan Terakhir', 'Pekerjaan', 'Status Perkawinan',
            'Golongan Darah', 'Dusun', 'Lorong', 'Alamat', 'RT', 'RW',
            'No. Telepon', 'No. Handphone', 'Email', 'No. KK'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_num, penduduk in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
            ws.cell(row=row_num, column=2, value=penduduk.nik).border = thin_border
            ws.cell(row=row_num, column=3, value=penduduk.name).border = thin_border
            ws.cell(row=row_num, column=4, value=penduduk.get_gender_display()).border = thin_border
            ws.cell(row=row_num, column=5, value=penduduk.birth_place or '').border = thin_border
            ws.cell(row=row_num, column=6, value=penduduk.birth_date.strftime('%d/%m/%Y') if penduduk.birth_date else '').border = thin_border
            ws.cell(row=row_num, column=7, value=penduduk.religion or '').border = thin_border
            ws.cell(row=row_num, column=8, value=penduduk.get_education_display() if penduduk.education else '').border = thin_border
            ws.cell(row=row_num, column=9, value=penduduk.occupation or '').border = thin_border
            ws.cell(row=row_num, column=10, value=penduduk.get_marital_status_display() if penduduk.marital_status else '').border = thin_border
            ws.cell(row=row_num, column=11, value=penduduk.blood_type or '').border = thin_border
            ws.cell(row=row_num, column=12, value=penduduk.dusun.name if penduduk.dusun else '').border = thin_border
            ws.cell(row=row_num, column=13, value=penduduk.lorong.nama_lorong if penduduk.lorong else '').border = thin_border
            ws.cell(row=row_num, column=14, value=penduduk.address or '').border = thin_border
            ws.cell(row=row_num, column=15, value=penduduk.rt_number or '').border = thin_border
            ws.cell(row=row_num, column=16, value=penduduk.rw_number or '').border = thin_border
            ws.cell(row=row_num, column=17, value=penduduk.phone_number or '').border = thin_border
            ws.cell(row=row_num, column=18, value=penduduk.mobile_number or '').border = thin_border
            ws.cell(row=row_num, column=19, value=penduduk.email or '').border = thin_border
            ws.cell(row=row_num, column=20, value=penduduk.kk_number or '').border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    def export_pdf(self, queryset):
        """Export data ke PDF dengan styling yang lebih baik"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Laporan_Data_Penduduk.pdf"'
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            rightMargin=50, 
            leftMargin=50, 
            topMargin=50, 
            bottomMargin=50
        )
        
        # Prepare data for table - Hanya 4 kolom penting saja
        data = [['NIK', 'Nama', 'Jenis Kelamin', 'Dusun']]
        
        for penduduk in queryset:  # Ambil semua data penduduk
            data.append([
                str(penduduk.nik),  # Pastikan NIK sebagai string
                penduduk.name[:30] + '...' if len(penduduk.name) > 30 else penduduk.name,
                penduduk.get_gender_display(),
                penduduk.dusun.name if penduduk.dusun else '-'
            ])
        
        # Create table with better column widths - 4 kolom saja
        col_widths = [1.8 * inch, 2.5 * inch, 1.2 * inch, 2.0 * inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Enhanced table styling
        table_style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),  # Professional blue
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),  # Light gray
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2E86AB')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ])
        
        table.setStyle(table_style)
        
        # Build PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Header with logo and title
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2E86AB'),
            fontName='Helvetica-Bold'
        )
        
        # Subtitle style
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#6C757D'),
            fontName='Helvetica'
        )
        
        # Info style
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=2,  # Right alignment
            textColor=colors.HexColor('#6C757D'),
            fontName='Helvetica'
        )
        
        # Add header
        elements.append(Paragraph("LAPORAN DATA PENDUDUK", header_style))
        elements.append(Paragraph("Desa Pulosarok", subtitle_style))
        elements.append(Paragraph(f"Dicetak pada: {pd.Timestamp.now().strftime('%d %B %Y, %H:%M WIB')}", info_style))
        elements.append(Spacer(1, 20))
        
        # Add table
        elements.append(table)
        
        # Add footer info
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#6C757D'),
            fontName='Helvetica'
        )
        elements.append(Paragraph(f"Total Data: {len(data)-1} | Halaman 1", footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF content
        buffer.seek(0)
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Laporan_Data_Penduduk.pdf"'
        
        return response
    
    def export_template(self):
        """Export template untuk import"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_penduduk.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Penduduk"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers dengan contoh data
        headers = [
            'NIK*', 'Nama Lengkap*', 'Jenis Kelamin*', 'Tempat Lahir*', 'Tanggal Lahir*',
            'Agama', 'Pendidikan Terakhir', 'Pekerjaan', 'Status Perkawinan',
            'Golongan Darah', 'Dusun*', 'Lorong', 'Alamat', 'RT', 'RW',
            'No. Telepon', 'No. Handphone', 'Email', 'No. KK'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Contoh data
        example_data = [
            '1234567890123456', 'John Doe', 'L', 'Jakarta', '01/01/1990',
            'ISLAM', 'SLTA', 'Karyawan', 'KAWIN', 'O',
            'Dusun 1', 'Lorong A', 'Jl. Contoh No. 1', '001', '001',
            '0211234567', '08123456789', 'john@example.com', '1234567890123456'
        ]
        
        for col_num, data in enumerate(example_data, 1):
            ws.cell(row=2, column=col_num, value=data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response


@method_decorator(csrf_exempt, name='dispatch')
class PendudukExportView(View, ExportImportMixin):
    """View untuk export data penduduk"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'excel')
        
        # Get filtered data
        queryset = Penduduk.objects.all()
        
        # Apply filters
        search_query = request.GET.get('search', '')
        dusun_filter = request.GET.get('dusun', '')
        gender_filter = request.GET.get('gender', '')
        
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | 
                Q(nik__icontains=search_query)
            )
        
        if dusun_filter:
            queryset = queryset.filter(dusun_id=dusun_filter)
        
        if gender_filter:
            queryset = queryset.filter(gender=gender_filter)
        
        return self.get_export_data(queryset, format_type)


@method_decorator(csrf_exempt, name='dispatch')
class PendudukImportView(View):
    """View untuk import data penduduk"""
    
    def post(self, request):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'Tidak ada file yang diupload'
                })
            
            file = request.FILES['file']
            skip_errors = request.POST.get('skip_errors', False)
            
            # Validasi file
            if not file.name.endswith(('.xlsx', '.csv')):
                return JsonResponse({
                    'success': False,
                    'error': 'Format file tidak didukung. Gunakan .xlsx atau .csv'
                })
            
            # Process file
            if file.name.endswith('.xlsx'):
                df = pd.read_excel(file)
            else:
                df = pd.read_csv(file)
            
            # Validasi kolom required
            required_columns = ['NIK', 'Nama Lengkap', 'Jenis Kelamin', 'Tempat Lahir', 'Tanggal Lahir', 'Dusun']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'success': False,
                    'error': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                })
            
            # Process data
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Get dusun
                    dusun_name = str(row['Dusun']).strip()
                    dusun = Dusun.objects.filter(name__icontains=dusun_name).first()
                    
                    if not dusun:
                        if not skip_errors:
                            errors.append(f"Baris {index + 2}: Dusun '{dusun_name}' tidak ditemukan")
                            error_count += 1
                            continue
                        else:
                            continue
                    
                    # Get lorong if provided
                    lorong = None
                    if 'Lorong' in df.columns and pd.notna(row['Lorong']):
                        lorong_name = str(row['Lorong']).strip()
                        lorong = Lorong.objects.filter(nama_lorong__icontains=lorong_name, dusun=dusun).first()
                    
                    # Validate required fields
                    if pd.isna(row['Tanggal Lahir']):
                        if not skip_errors:
                            errors.append(f"Baris {index + 2}: Tanggal Lahir harus diisi")
                            error_count += 1
                            continue
                        else:
                            continue
                    
                    # Create penduduk with proper field mapping
                    # Map gender to model choices
                    gender_value = str(row['Jenis Kelamin']).strip().upper()
                    gender_mapping = {
                        'LAKI-LAKI': 'L',
                        'L': 'L',
                        'PEREMPUAN': 'P',
                        'P': 'P'
                    }
                    mapped_gender = gender_mapping.get(gender_value, 'L')
                    
                    penduduk_data = {
                        'nik': str(row['NIK']).strip(),
                        'name': str(row['Nama Lengkap']).strip(),
                        'gender': mapped_gender,
                        'birth_place': str(row['Tempat Lahir']).strip(),
                        'birth_date': pd.to_datetime(row['Tanggal Lahir']).date(),
                        'dusun': dusun,
                        'lorong': lorong,
                        'religion': 'Islam',  # Default value
                        'marital_status': 'BELUM_KAWIN',  # Default value
                        'address': str(row.get('Alamat', 'Alamat tidak diisi')).strip() if pd.notna(row.get('Alamat', '')) else 'Alamat tidak diisi',
                    }
                    
                    # Optional fields with proper mapping
                    if 'Agama' in df.columns and pd.notna(row['Agama']):
                        agama_value = str(row['Agama']).strip().upper()
                        # Map agama values to model choices
                        agama_mapping = {
                            'ISLAM': 'Islam',
                            'KRISTEN': 'Kristen Protestan',
                            'KATOLIK': 'Kristen Katolik',
                            'HINDU': 'Hindu',
                            'BUDDHA': 'Buddha',
                            'KONGHUCU': 'Konghucu'
                        }
                        penduduk_data['religion'] = agama_mapping.get(agama_value, 'Islam')
                    
                    if 'Pendidikan Terakhir' in df.columns and pd.notna(row['Pendidikan Terakhir']):
                        pendidikan_value = str(row['Pendidikan Terakhir']).strip().upper()
                        # Map pendidikan values to model choices
                        pendidikan_mapping = {
                            'TIDAK SEKOLAH': 'TIDAK_BELUM_SEKOLAH',
                            'SD': 'TAMAT_SD',
                            'SMP': 'SLTP',
                            'SMA': 'SLTA',
                            'D3': 'D3',
                            'S1': 'D4_S1',
                            'S2': 'S2',
                            'S3': 'S3'
                        }
                        penduduk_data['education'] = pendidikan_mapping.get(pendidikan_value, 'TIDAK_BELUM_SEKOLAH')
                    
                    if 'Pekerjaan' in df.columns and pd.notna(row['Pekerjaan']):
                        penduduk_data['occupation'] = str(row['Pekerjaan']).strip()
                    
                    if 'Status Perkawinan' in df.columns and pd.notna(row['Status Perkawinan']):
                        status_value = str(row['Status Perkawinan']).strip().upper()
                        # Map status perkawinan values to model choices
                        status_mapping = {
                            'BELUM KAWIN': 'BELUM_KAWIN',
                            'KAWIN': 'KAWIN',
                            'CERAI HIDUP': 'CERAI_HIDUP',
                            'CERAI MATI': 'CERAI_MATI'
                        }
                        penduduk_data['marital_status'] = status_mapping.get(status_value, 'BELUM_KAWIN')
                    
                    if 'Golongan Darah' in df.columns and pd.notna(row['Golongan Darah']):
                        penduduk_data['blood_type'] = str(row['Golongan Darah']).strip().upper()
                    
                    if 'RT' in df.columns and pd.notna(row['RT']):
                        rt_number = str(row['RT']).strip()
                        # Find RT object
                        rt_obj = RT.objects.filter(rt_number=rt_number, dusun=dusun).first()
                        if rt_obj:
                            penduduk_data['rt'] = rt_obj
                    
                    if 'RW' in df.columns and pd.notna(row['RW']):
                        rw_number = str(row['RW']).strip()
                        # Find RW object
                        rw_obj = RW.objects.filter(rw_number=rw_number).first()
                        if rw_obj:
                            penduduk_data['rw'] = rw_obj
                    
                    if 'No. Telepon' in df.columns and pd.notna(row['No. Telepon']):
                        penduduk_data['phone_number'] = str(row['No. Telepon']).strip()
                    
                    if 'No. Handphone' in df.columns and pd.notna(row['No. Handphone']):
                        penduduk_data['mobile_number'] = str(row['No. Handphone']).strip()
                    
                    if 'Email' in df.columns and pd.notna(row['Email']):
                        penduduk_data['email'] = str(row['Email']).strip()
                    
                    if 'No. KK' in df.columns and pd.notna(row['No. KK']):
                        penduduk_data['kk_number'] = str(row['No. KK']).strip()
                    
                    # Check if penduduk already exists
                    if Penduduk.objects.filter(nik=penduduk_data['nik']).exists():
                        if not skip_errors:
                            errors.append(f"Baris {index + 2}: NIK {penduduk_data['nik']} sudah ada")
                            error_count += 1
                            continue
                        else:
                            continue
                    
                    # Create penduduk object
                    try:
                        penduduk = Penduduk.objects.create(**penduduk_data)
                        success_count += 1
                        print(f"Successfully created penduduk: {penduduk.name} (NIK: {penduduk.nik})")
                    except Exception as create_error:
                        print(f"Error creating penduduk: {str(create_error)}")
                        print(f"Penduduk data: {penduduk_data}")
                        if not skip_errors:
                            errors.append(f"Baris {index + 2}: Error creating penduduk - {str(create_error)}")
                            error_count += 1
                        else:
                            error_count += 1
                    
                except Exception as e:
                    if not skip_errors:
                        errors.append(f"Baris {index + 2}: {str(e)}")
                        error_count += 1
                    else:
                        error_count += 1
            
            message = f"Import berhasil! {success_count} data berhasil diimport"
            if error_count > 0:
                message += f", {error_count} data error"
            
            return JsonResponse({
                'success': True,
                'message': message,
                'success_count': int(success_count),
                'error_count': int(error_count),
                'errors': errors[:10]  # Limit errors to first 10
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Terjadi kesalahan: {str(e)}'
            })

def quick_import(request, model_name):
    """Quick import function for different models"""
    if model_name == 'penduduk':
        return PendudukImportView.as_view()(request)
    elif model_name == 'rt':
        return RTImportView.as_view()(request)
    elif model_name == 'rw':
        return RWImportView.as_view()(request)
    elif model_name == 'dusun':
        return DusunImportView.as_view()(request)
    elif model_name == 'lorong':
        return LorongImportView.as_view()(request)
    elif model_name == 'keluarga':
        return KeluargaImportView.as_view()(request)
    else:
        return JsonResponse({
            'success': False,
            'error': f'Model {model_name} tidak didukung untuk import'
        })


class RTImportView(View):
    """Import view untuk RT"""
    
    def post(self, request):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'File tidak ditemukan'
                })
            
            file = request.FILES['file']
            
            # Cek format file
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Format file tidak didukung. Gunakan Excel (.xlsx atau .xls)'
                })
            
            # Baca file Excel
            df = pd.read_excel(file)
            
            # Validasi kolom yang diperlukan
            required_columns = ['RW_ID', 'RT_Number', 'Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'success': False,
                    'error': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                })
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Cari RW berdasarkan ID
                    rw = RW.objects.get(id=row['RW_ID'])
                    
                    # Buat RT baru
                    rt = RT.objects.create(
                        rw=rw,
                        rt_number=str(row['RT_Number']).strip(),
                        description=str(row.get('Description', '')).strip() if pd.notna(row.get('Description', '')) else '',
                        is_active=True
                    )
                    
                    success_count += 1
                    print(f"Successfully created RT: {rt.rt_number} in RW {rt.rw.rw_number}")
                    
                except RW.DoesNotExist:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: RW dengan ID {row['RW_ID']} tidak ditemukan")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Import selesai dengan {success_count} data berhasil, {error_count} data gagal',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # Batasi error yang ditampilkan
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Terjadi kesalahan: {str(e)}'
            })


class RWImportView(View):
    """Import view untuk RW"""
    
    def post(self, request):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'File tidak ditemukan'
                })
            
            file = request.FILES['file']
            
            # Cek format file
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Format file tidak didukung. Gunakan Excel (.xlsx atau .xls)'
                })
            
            # Baca file Excel
            df = pd.read_excel(file)
            
            # Validasi kolom yang diperlukan
            required_columns = ['Dusun_ID', 'RW_Number', 'Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'success': False,
                    'error': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                })
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Cari Dusun berdasarkan ID
                    dusun = Dusun.objects.get(id=row['Dusun_ID'])
                    
                    # Buat RW baru
                    rw = RW.objects.create(
                        dusun=dusun,
                        rw_number=str(row['RW_Number']).strip(),
                        description=str(row.get('Description', '')).strip() if pd.notna(row.get('Description', '')) else '',
                        is_active=True
                    )
                    
                    success_count += 1
                    print(f"Successfully created RW: {rw.rw_number} in {rw.dusun.name}")
                    
                except Dusun.DoesNotExist:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: Dusun dengan ID {row['Dusun_ID']} tidak ditemukan")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Import selesai dengan {success_count} data berhasil, {error_count} data gagal',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # Batasi error yang ditampilkan
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Terjadi kesalahan: {str(e)}'
            })


class DusunImportView(View):
    """Import view untuk Dusun"""
    
    def post(self, request):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'File tidak ditemukan'
                })
            
            file = request.FILES['file']
            
            # Cek format file
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Format file tidak didukung. Gunakan Excel (.xlsx atau .xls)'
                })
            
            # Baca file Excel
            df = pd.read_excel(file)
            
            # Validasi kolom yang diperlukan
            required_columns = ['Name', 'Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'success': False,
                    'error': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                })
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Buat Dusun baru
                    dusun = Dusun.objects.create(
                        name=str(row['Name']).strip(),
                        description=str(row.get('Description', '')).strip() if pd.notna(row.get('Description', '')) else '',
                        is_active=True
                    )
                    
                    success_count += 1
                    print(f"Successfully created Dusun: {dusun.name}")
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Import selesai dengan {success_count} data berhasil, {error_count} data gagal',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # Batasi error yang ditampilkan
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Terjadi kesalahan: {str(e)}'
            })


class LorongImportView(View):
    """Import view untuk Lorong"""
    
    def post(self, request):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'File tidak ditemukan'
                })
            
            file = request.FILES['file']
            
            # Cek format file
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Format file tidak didukung. Gunakan Excel (.xlsx atau .xls)'
                })
            
            # Baca file Excel
            df = pd.read_excel(file)
            
            # Validasi kolom yang diperlukan
            required_columns = ['Dusun_ID', 'Lorong_Code', 'Name', 'Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'success': False,
                    'error': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                })
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Cari Dusun berdasarkan ID
                    dusun = Dusun.objects.get(id=row['Dusun_ID'])
                    
                    # Buat Lorong baru
                    lorong = Lorong.objects.create(
                        dusun=dusun,
                        lorong_code=str(row['Lorong_Code']).strip(),
                        name=str(row['Name']).strip(),
                        description=str(row.get('Description', '')).strip() if pd.notna(row.get('Description', '')) else '',
                        is_active=True
                    )
                    
                    success_count += 1
                    print(f"Successfully created Lorong: {lorong.name} in {lorong.dusun.name}")
                    
                except Dusun.DoesNotExist:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: Dusun dengan ID {row['Dusun_ID']} tidak ditemukan")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Import selesai dengan {success_count} data berhasil, {error_count} data gagal',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # Batasi error yang ditampilkan
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Terjadi kesalahan: {str(e)}'
            })


class KeluargaImportView(View):
    """Import view untuk Keluarga"""
    
    def post(self, request):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'File tidak ditemukan'
                })
            
            file = request.FILES['file']
            
            # Cek format file
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Format file tidak didukung. Gunakan Excel (.xlsx atau .xls)'
                })
            
            # Baca file Excel
            df = pd.read_excel(file)
            
            # Validasi kolom yang diperlukan
            required_columns = ['Family_Name', 'Address', 'RT_ID']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'success': False,
                    'error': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                })
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Cari RT berdasarkan ID
                    rt = RT.objects.get(id=row['RT_ID'])
                    
                    # Buat Keluarga baru
                    keluarga = Keluarga.objects.create(
                        family_name=str(row['Family_Name']).strip(),
                        address=str(row['Address']).strip(),
                        rt=rt,
                        is_active=True
                    )
                    
                    success_count += 1
                    print(f"Successfully created Keluarga: {keluarga.family_name}")
                    
                except RT.DoesNotExist:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: RT dengan ID {row['RT_ID']} tidak ditemukan")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Baris {index + 2}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Import selesai dengan {success_count} data berhasil, {error_count} data gagal',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # Batasi error yang ditampilkan
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Terjadi kesalahan: {str(e)}'
            })
