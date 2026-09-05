"""
Views untuk Import/Export Data References
Menyediakan interface web untuk operasi import/export
"""

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.db import transaction
import os
import json
import pandas as pd
from datetime import datetime
import logging
from openpyxl.utils import get_column_letter

from .import_export import ImportExportManager
from .models import Penduduk, Family, Dusun, Lorong, DisabilitasData, DisabilitasType, ReligionReference, Keluarga, Pelajar, RW, RT

logger = logging.getLogger(__name__)

@login_required
def import_export_dashboard(request):
    """Dashboard utama untuk operasi import/export"""
    context = {
        'title': 'Import/Export Data',
        'models': [
            {'name': 'penduduk', 'verbose_name': 'Data Penduduk', 'count': Penduduk.objects.count()},
            {'name': 'keluarga', 'verbose_name': 'Data Keluarga', 'count': Keluarga.objects.count()},
            {'name': 'family', 'verbose_name': 'Data Family', 'count': Family.objects.count()},
            {'name': 'dusun', 'verbose_name': 'Data Dusun', 'count': Dusun.objects.count()},
            {'name': 'lorong', 'verbose_name': 'Data Lorong', 'count': Lorong.objects.count()},
            {'name': 'pelajar', 'verbose_name': 'Data Pelajar', 'count': Pelajar.objects.count()},
            {'name': 'disabilitas', 'verbose_name': 'Data Disabilitas', 'count': DisabilitasData.objects.count()},
            {'name': 'disabilitas_type', 'verbose_name': 'Jenis Disabilitas', 'count': DisabilitasType.objects.count()},
            {'name': 'religion', 'verbose_name': 'Data Agama', 'count': ReligionReference.objects.count()},
            {'name': 'rw', 'verbose_name': 'Data RW', 'count': RW.objects.count()},
            {'name': 'rt', 'verbose_name': 'Data RT', 'count': RT.objects.count()},
        ],
        'formats': ['excel', 'csv', 'json', 'pdf'],
    }
    return render(request, 'admin_panel/references/import_export_dashboard.html', context)

@login_required
@require_http_methods(["GET", "POST"])
def export_data(request, model_name, format=None):
    """Export data dalam format yang ditentukan - langsung download"""
    try:
        # Langsung export tanpa form
        format_type = format or 'excel'
        include_related = True
        
        # Dapatkan filter dari query parameters
        filters = {}
        if request.GET.get('is_active'):
            filters['is_active'] = request.GET.get('is_active') == 'true'
        if request.GET.get('date_from'):
            filters['date_from'] = request.GET.get('date_from')
        if request.GET.get('date_to'):
            filters['date_to'] = request.GET.get('date_to')
        if request.GET.get('dusun'):
            filters['dusun'] = request.GET.get('dusun')
        if request.GET.get('gender'):
            filters['gender'] = request.GET.get('gender')
        if request.GET.get('religion'):
            filters['religion'] = request.GET.get('religion')
        
        manager = ImportExportManager()
        response = manager.export_data(model_name, format_type, filters, include_related)
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error saat mengekspor data: {str(e)}'
        })

@login_required
@require_http_methods(["GET", "POST"])
def import_data(request, model_name):
    """Import data dari file"""
    if request.method == 'GET':
        # Tampilkan form import
        context = {
            'title': f'Import Data {model_name.title()}',
            'model_name': model_name,
            'formats': ['excel', 'csv', 'json'],
        }
        return render(request, 'admin_panel/references/import_form.html', context)
    
    else:
        # Handle import request
        format_type = request.POST.get('format', 'excel')
        validate_only = request.POST.get('validate_only', 'off') == 'on'
        
        if 'file' not in request.FILES:
            messages.error(request, 'Tidak ada file yang dipilih')
            return redirect('admin_panel:import_data', model_name=model_name)
        
        file = request.FILES['file']
        
        # Simpan file yang diupload
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp', f'{model_name}_import_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format_type}')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        try:
            manager = ImportExportManager()
            result = manager.import_data(model_name, file_path, format_type, validate_only)
            
            if result['success']:
                if validate_only:
                    messages.success(request, f'Validasi berhasil! {result["total_rows"]} baris data valid.')
                else:
                    messages.success(request, f'Import berhasil! {result["imported"]} data berhasil diimpor dari {result["total_rows"]} baris.')
            else:
                messages.warning(request, f'Import selesai dengan {len(result["errors"])} error.')
                for error in result['errors'][:5]:  # Tampilkan 5 error pertama
                    messages.error(request, f'Baris {error["row"]}: {error["error"]}')
            
            # Bersihkan file temp
            if os.path.exists(file_path):
                os.remove(file_path)
            
            return redirect('admin_panel:import_export_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error saat mengimpor data: {str(e)}')
            # Bersihkan file temp
            if os.path.exists(file_path):
                os.remove(file_path)
            return redirect('admin_panel:import_data', model_name=model_name)

@login_required
def download_template(request, model_name, format_type='excel'):
    """Download template untuk import data"""
    try:
        manager = ImportExportManager()
        response = manager.create_template(model_name, format_type)
        return response
        
    except Exception as e:
        messages.error(request, f'Error saat mengunduh template: {str(e)}')
        return redirect('admin_panel:import_export_dashboard')

@login_required
def bulk_export(request):
    """Bulk export semua jenis data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        format_type = request.POST.get('format', 'excel')
        models = request.POST.getlist('models')
        
        if not models:
            return JsonResponse({'error': 'No models selected'}, status=400)
        
        # Buat zip file dengan semua export
        import zipfile
        import io
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            manager = ImportExportManager()
            
            for model_name in models:
                try:
                    response = manager.export_data(model_name, format_type, {}, True)
                    content = response.content
                    
                    filename = f"{model_name}_export.{format_type}"
                    zip_file.writestr(filename, content)
                    
                except Exception as e:
                    # Tambahkan file error ke zip
                    error_content = f"Error exporting {model_name}: {str(e)}"
                    zip_file.writestr(f"{model_name}_error.txt", error_content)
        
        zip_buffer.seek(0)
        
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="bulk_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def export_statistics(request):
    """Export statistik komprehensif"""
    try:
        from .import_export import ImportExportManager
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # Hapus sheet default
        wb.remove(wb.active)
        
        # Buat sheet ringkasan
        ws_summary = wb.create_sheet("Ringkasan Statistik")
        
        # Dapatkan statistik
        stats = {
            'Total Penduduk': Penduduk.objects.filter(is_active=True).count(),
            'Total Keluarga': Keluarga.objects.filter(is_active=True).count(),
            'Total Family': Family.objects.filter(is_active=True).count(),
            'Total Dusun': Dusun.objects.filter(is_active=True).count(),
            'Total Lorong': Lorong.objects.filter(is_active=True).count(),
            'Total Disabilitas': DisabilitasData.objects.filter(is_active=True).count(),
            'Total Agama': ReligionReference.objects.filter(is_active=True).count(),
        }
        
        # Tulis ringkasan
        ws_summary['A1'] = 'STATISTIK KEPENDUDUKAN DESA PULOSAROK'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_summary['A1'].font = Font(bold=True, color="FFFFFF")
        
        row = 3
        for key, value in stats.items():
            ws_summary[f'A{row}'] = key
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Tambahkan breakdown demografi
        ws_demo = wb.create_sheet("Demografi")
        
        # Distribusi gender
        male_count = Penduduk.objects.filter(is_active=True, gender='L').count()
        female_count = Penduduk.objects.filter(is_active=True, gender='P').count()
        
        ws_demo['A1'] = 'Distribusi Jenis Kelamin'
        ws_demo['A1'].font = Font(bold=True)
        ws_demo['A2'] = 'Laki-laki'
        ws_demo['B2'] = male_count
        ws_demo['A3'] = 'Perempuan'
        ws_demo['B3'] = female_count
        
        # Distribusi agama
        ws_demo['A5'] = 'Distribusi Agama'
        ws_demo['A5'].font = Font(bold=True)
        
        row = 6
        for religion in Penduduk.objects.filter(is_active=True).values_list('religion', flat=True).distinct():
            count = Penduduk.objects.filter(is_active=True, religion=religion).count()
            ws_demo[f'A{row}'] = religion
            ws_demo[f'B{row}'] = count
            row += 1
        
        # Distribusi dusun
        ws_dusun = wb.create_sheet("Distribusi Dusun")
        
        ws_dusun['A1'] = 'Dusun'
        ws_dusun['B1'] = 'Jumlah Penduduk'
        ws_dusun['C1'] = 'Luas Area (Ha)'
        ws_dusun['D1'] = 'Kepadatan (jiwa/ha)'
        
        # Style headers
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_dusun[col].font = Font(bold=True)
            ws_dusun[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws_dusun[col].font = Font(bold=True, color="FFFFFF")
        
        row = 2
        for dusun in Dusun.objects.filter(is_active=True):
            pop_count = Penduduk.objects.filter(dusun=dusun, is_active=True).count()
            density = pop_count / float(dusun.area_size) if dusun.area_size else 0
            
            ws_dusun[f'A{row}'] = dusun.name
            ws_dusun[f'B{row}'] = pop_count
            ws_dusun[f'C{row}'] = dusun.area_size or 0
            ws_dusun[f'D{row}'] = round(density, 2)
            row += 1
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="statistik_kependudukan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error saat mengekspor statistik: {str(e)}')
        return redirect('admin_panel:import_export_dashboard')

@login_required
@csrf_exempt
def api_import_preview(request, model_name):
    """API endpoint untuk preview import"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        # Simpan file temp
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp', f'preview_{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{file_extension}')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        # Baca dan preview data
        manager = ImportExportManager()
        
        if file_extension in ['xlsx', 'xls']:
            import pandas as pd
            df = pd.read_excel(file_path)
            preview_data = df.head(10).to_dict('records')
            columns = list(df.columns)
        elif file_extension == 'csv':
            import pandas as pd
            df = pd.read_csv(file_path, encoding='utf-8')
            preview_data = df.head(10).to_dict('records')
            columns = list(df.columns)
        elif file_extension == 'json':
            import json
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            preview_data = data[:10]
            columns = list(data[0].keys()) if data else []
        else:
            return JsonResponse({'error': f'Unsupported file format: {file_extension}'}, status=400)
        
        # Bersihkan file temp
        if os.path.exists(file_path):
            os.remove(file_path)
        
        return JsonResponse({
            'success': True,
            'preview_data': preview_data,
            'columns': columns,
            'total_rows': len(preview_data)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Individual Import/Export Functions untuk setiap model
@login_required
def penduduk_import_export(request):
    """Import/Export khusus untuk penduduk"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'import':
            return handle_import(request, 'penduduk')
        elif action == 'export':
            return handle_export(request, 'penduduk')
    
    context = {
        'title': 'Import/Export Data Penduduk',
        'model_name': 'penduduk',
        'formats': ['excel', 'csv', 'json'],
    }
    return render(request, 'admin_panel/references/import_export_form.html', context)

@login_required
def dusun_import_export(request):
    """Import/Export khusus untuk dusun"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'import':
            return handle_import(request, 'dusun')
        elif action == 'export':
            return handle_export(request, 'dusun')
    
    context = {
        'title': 'Import/Export Data Dusun',
        'model_name': 'dusun',
        'formats': ['excel', 'csv', 'json'],
    }
    return render(request, 'admin_panel/references/import_export_form.html', context)

def handle_import(request, model_name):
    """Handle import request"""
    format_type = request.POST.get('format', 'excel')
    validate_only = request.POST.get('validate_only', 'off') == 'on'
    
    if 'file' not in request.FILES:
        messages.error(request, 'Tidak ada file yang dipilih')
        return redirect('admin_panel:import_export_dashboard')
    
    file = request.FILES['file']
    
    # Simpan file yang diupload
    file_path = os.path.join(settings.MEDIA_ROOT, 'temp', f'{model_name}_import_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format_type}')
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    with open(file_path, 'wb') as destination:
        for chunk in file.chunks():
            destination.write(chunk)
    
    try:
        manager = ImportExportManager()
        result = manager.import_data(model_name, file_path, format_type, validate_only)
        
        if result['success']:
            if validate_only:
                messages.success(request, f'Validasi berhasil! {result["total_rows"]} baris data valid.')
            else:
                messages.success(request, f'Import berhasil! {result["imported"]} data berhasil diimpor dari {result["total_rows"]} baris.')
        else:
            messages.warning(request, f'Import selesai dengan {len(result["errors"])} error.')
            for error in result['errors'][:5]:
                messages.error(request, f'Baris {error["row"]}: {error["error"]}')
        
        # Bersihkan file temp
        if os.path.exists(file_path):
            os.remove(file_path)
        
        return redirect('admin_panel:import_export_dashboard')
        
    except Exception as e:
        messages.error(request, f'Error saat mengimpor data: {str(e)}')
        if os.path.exists(file_path):
            os.remove(file_path)
        return redirect('admin_panel:import_export_dashboard')

def handle_export(request, model_name):
    """Handle export request"""
    format_type = request.POST.get('format', 'excel')
    include_related = request.POST.get('include_related', 'on') == 'on'
    
    # Dapatkan filter
    filters = {}
    if request.POST.get('is_active'):
        filters['is_active'] = request.POST.get('is_active') == 'true'
    if request.POST.get('date_from'):
        filters['date_from'] = request.POST.get('date_from')
    if request.POST.get('date_to'):
        filters['date_to'] = request.POST.get('date_to')
    if request.POST.get('dusun'):
        filters['dusun'] = request.POST.get('dusun')
    if request.POST.get('gender'):
        filters['gender'] = request.POST.get('gender')
    if request.POST.get('religion'):
        filters['religion'] = request.POST.get('religion')
    
    try:
        manager = ImportExportManager()
        response = manager.export_data(model_name, format_type, filters, include_related)
        
        messages.success(request, f'Data {model_name} berhasil diekspor dalam format {format_type.upper()}')
        return response
        
    except Exception as e:
        messages.error(request, f'Error saat mengekspor data: {str(e)}')
        return redirect('admin_panel:import_export_dashboard')

# New enhanced import/export functions
@login_required
@csrf_exempt
@require_http_methods(["POST"])
def quick_export(request, model_name):
    """Quick export dengan format default (Excel)"""
    try:
        format_type = request.POST.get('format', 'excel')
        manager = ImportExportManager()
        response = manager.export_data(model_name, format_type, {}, True)
        return response
    except Exception as e:
        logger.error(f"Quick export error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def quick_import(request, model_name):
    """Quick import dengan validasi otomatis dan smart detection untuk penduduk"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        format_type = file.name.split('.')[-1].lower()
        
        # Simpan file temporary
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp', f'{model_name}_import_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format_type}')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        # Import data dengan smart detection untuk penduduk
        manager = ImportExportManager()
        result = manager.import_data(model_name, file_path, format_type, False)
        
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
        
        if result['success']:
            # Tambahkan info auto-create untuk penduduk
            auto_created_info = ""
            if model_name == 'penduduk' and result.get('auto_created'):
                auto_created_info = f" Auto-created: {result['auto_created']}"
            
            return JsonResponse({
                'success': True,
                'message': f'Import berhasil! {result["imported"]} data berhasil diimpor dari {result["total_rows"]} baris.{auto_created_info}',
                'imported': result['imported'],
                'total_rows': result['total_rows'],
                'auto_created': result.get('auto_created', {})
            })
        else:
            error_details = []
            for error in result['errors'][:5]:  # Limit to first 5 errors
                error_details.append(f"Baris {error['row']}: {error['error']}")
            
            return JsonResponse({
                'success': False,
                'message': f'Import selesai dengan {len(result["errors"])} error.',
                'errors': result['errors'][:10],  # Limit to first 10 errors
                'error_details': error_details
            })
            
    except Exception as e:
        logger.error(f"Quick import error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["GET"])
def get_import_template(request, model_name):
    """Get import template untuk model tertentu"""
    try:
        format_type = request.GET.get('format', 'excel')
        manager = ImportExportManager()
        response = manager.create_template(model_name, format_type)
        return response
    except Exception as e:
        logger.error(f"Template generation error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def bulk_export_selected(request):
    """Export data yang dipilih dari list view"""
    try:
        selected_ids = request.POST.getlist('selected_ids[]')
        model_name = request.POST.get('model_name')
        format_type = request.POST.get('format', 'excel')
        
        if not selected_ids:
            return JsonResponse({'error': 'No data selected'}, status=400)
        
        # Get model class
        model_mapping = {
            'penduduk': Penduduk,
            'keluarga': Keluarga,
            'family': Family,
            'dusun': Dusun,
            'lorong': Lorong,
            'pelajar': Pelajar,
            'disabilitas': DisabilitasData,
            'disabilitas_type': DisabilitasType,
            'religion': ReligionReference,
            'rw': RW,
            'rt': RT,
        }
        
        if model_name not in model_mapping:
            return JsonResponse({'error': 'Invalid model name'}, status=400)
        
        model_class = model_mapping[model_name]
        queryset = model_class.objects.filter(id__in=selected_ids)
        
        manager = ImportExportManager()
        
        if format_type == 'excel':
            return manager._export_to_excel(model_name, queryset, True)
        elif format_type == 'csv':
            return manager._export_to_csv(model_name, queryset, True)
        elif format_type == 'json':
            return manager._export_to_json(model_name, queryset, True)
        else:
            return JsonResponse({'error': 'Unsupported format'}, status=400)
            
    except Exception as e:
        logger.error(f"Bulk export error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def validate_import_file(request):
    """Validate import file sebelum import"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        model_name = request.POST.get('model_name')
        format_type = file.name.split('.')[-1].lower()
        
        # Simpan file temporary
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp', f'validate_{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format_type}')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        # Validate data
        manager = ImportExportManager()
        result = manager.import_data(model_name, file_path, format_type, True)
        
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
        
        return JsonResponse({
            'success': result['success'],
            'total_rows': result['total_rows'],
            'valid_rows': result['total_rows'] - len(result['errors']),
            'error_count': len(result['errors']),
            'errors': result['errors'][:10]  # Limit to first 10 errors
        })
        
    except Exception as e:
        logger.error(f"Validation error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def bulk_action(request):
    """Handle bulk actions (delete, activate, deactivate)"""
    try:
        model_name = request.POST.get('model')
        action = request.POST.get('action')
        items = request.POST.get('items', '').split(',')
        
        if not model_name or not action or not items:
            return JsonResponse({
                'success': False,
                'message': 'Parameter tidak lengkap'
            }, status=400)
        
        # Get model class
        model_map = {
            'dusun': Dusun,
            'lorong': Lorong,
            'penduduk': Penduduk,
            'keluarga': Keluarga,
            'pelajar': Pelajar,
            'disabilitas': DisabilitasData,
            'rt': RT,
            'rw': RW
        }
        
        if model_name not in model_map:
            return JsonResponse({
                'success': False,
                'message': 'Model tidak valid'
            }, status=400)
        
        model_class = model_map[model_name]
        
        # Filter valid IDs
        valid_ids = []
        for item_id in items:
            try:
                valid_ids.append(int(item_id))
            except (ValueError, TypeError):
                continue
        
        if not valid_ids:
            return JsonResponse({
                'success': False,
                'message': 'Tidak ada item valid yang dipilih'
            }, status=400)
        
        # Perform bulk action
        with transaction.atomic():
            if action == 'delete':
                deleted_count = model_class.objects.filter(id__in=valid_ids).delete()[0]
                message = f'Berhasil menghapus {deleted_count} item'
                
            elif action == 'activate':
                updated_count = model_class.objects.filter(id__in=valid_ids).update(is_active=True)
                message = f'Berhasil mengaktifkan {updated_count} item'
                
            elif action == 'deactivate':
                updated_count = model_class.objects.filter(id__in=valid_ids).update(is_active=False)
                message = f'Berhasil menonaktifkan {updated_count} item'
                
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Aksi tidak valid'
                }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        logger.error(f"Bulk action error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

