from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.db.models import Count, Q, Case, When, IntegerField, Avg, Sum
from django.utils import timezone
from datetime import datetime, date, timedelta
from .models import Penduduk, Dusun, Family, DisabilitasData, DisabilitasType, Lorong, Pelajar, ReligionReference, Keluarga, RW, RT
from .forms import PendudukForm, DusunForm, LorongForm, RWForm, RTForm
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_page
from django.contrib import messages
import json

# Import other app models for comprehensive statistics
try:
    from news.models import News
    NEWS_AVAILABLE = True
except ImportError:
    NEWS_AVAILABLE = False

# try:
#     from business.models import UMKM
#     BUSINESS_AVAILABLE = True
# except ImportError:
BUSINESS_AVAILABLE = False

try:
    from tourism.models import TourismLocation
    TOURISM_AVAILABLE = True
except ImportError:
    TOURISM_AVAILABLE = False


def population_page(request):
    """Main population statistics page"""
    context = {
        'page_title': 'Data Penduduk Desa Pulosarok',
        'meta_description': 'Statistik dan demografi penduduk Desa Pulosarok'
    }
    return render(request, 'public/references/population.html', context)


@login_required
def dusun_list(request):
    """View untuk menampilkan daftar dusun"""
    dusuns = Dusun.objects.filter(is_active=True).order_by('name')
    context = {
        'dusuns': dusuns
    }
    return render(request, 'admin_panel/references/dusun_list.html', context)

@login_required
def dusun_add(request):
    """View untuk menambah dusun baru"""
    if request.method == 'POST':
        try:
            # Handle both JSON and form data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST
            
            kepala_dusun_id = data.get('kepala_dusun')
            kepala_dusun = None
            if kepala_dusun_id:
                kepala_dusun = Penduduk.objects.get(id=kepala_dusun_id)
            
            # Handle area_size conversion
            area_size = data.get('area_size')
            if area_size:
                try:
                    # Replace comma with dot for decimal parsing
                    area_size = float(str(area_size).replace(',', '.'))
                except (ValueError, TypeError):
                    area_size = None
            
            dusun = Dusun.objects.create(
                name=data['name'],
                code=data['code'],
                description=data.get('description', ''),
                area_size=area_size,
                kepala_dusun=kepala_dusun,
                is_active=True
            )
            
            if request.content_type == 'application/json':
                return JsonResponse({
                    'success': True,
                    'data': {
                        'id': dusun.id,
                        'name': dusun.name,
                        'code': dusun.code
                    }
                })
            else:
                messages.success(request, 'Data dusun berhasil ditambahkan.')
                return redirect('admin_panel:references_dusun_list')
        except Exception as e:
            if request.content_type == 'application/json':
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
            else:
                messages.error(request, f'Error: {str(e)}')
                return redirect('admin_panel:references_dusun_list')
    
    # Get all dusun for selection
    dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
    context = {
        'dusun_list': dusun_list,
        'form_title': 'Tambah Dusun',
        'form_icon': 'plus',
        'form_subtitle': 'Lengkapi data dusun baru'
    }
    return render(request, 'admin_panel/references/create/dusun_form.html', context)

@login_required
def dusun_detail(request, dusun_id):
    """View untuk menampilkan detail dusun"""
    try:
        dusun = get_object_or_404(Dusun, id=dusun_id)
        context = {
            'dusun': dusun
        }
        return render(request, 'admin_panel/references/detail/dusun_detail.html', context)
    except Dusun.DoesNotExist:
        messages.error(request, 'Data dusun tidak ditemukan.')
        return redirect('admin_panel:references_dusun_list')

@login_required
def dusun_edit(request, dusun_id):
    """View untuk mengedit dusun"""
    try:
        dusun = get_object_or_404(Dusun, id=dusun_id)
        
        if request.method == 'POST':
            # Handle form submission
            dusun.name = request.POST.get('name', dusun.name)
            dusun.code = request.POST.get('code', dusun.code)
            dusun.description = request.POST.get('description', dusun.description)
            
            # Handle area_size
            area_size = request.POST.get('area_size')
            if area_size:
                try:
                    dusun.area_size = float(area_size)
                except ValueError:
                    pass
            
            # Handle kepala_dusun
            kepala_dusun_id = request.POST.get('kepala_dusun')
            if kepala_dusun_id:
                try:
                    dusun.kepala_dusun = Penduduk.objects.get(id=kepala_dusun_id)
                except Penduduk.DoesNotExist:
                    dusun.kepala_dusun = None
            else:
                dusun.kepala_dusun = None
            
            # Handle is_active
            dusun.is_active = 'is_active' in request.POST
            
            # Population count will be auto-calculated in model save()
            dusun.save()
            messages.success(request, 'Data dusun berhasil diperbarui.')
            return redirect('admin_panel:references_dusun_list')
        
        # Get all dusun for selection
        dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
        context = {
            'dusun': dusun,
            'dusun_list': dusun_list,
            'form_title': 'Edit Dusun',
            'form_icon': 'edit',
            'form_subtitle': 'Perbarui data dusun yang sudah ada'
        }
        return render(request, 'admin_panel/references/create/dusun_form.html', context)
    except Dusun.DoesNotExist:
        messages.error(request, 'Data dusun tidak ditemukan.')
        return redirect('admin_panel:references_dusun_list')

@login_required
def dusun_update(request, dusun_id):
    """View untuk mengupdate dusun"""
    try:
        dusun = get_object_or_404(Dusun, id=dusun_id)
        if request.method == 'POST':
            data = json.loads(request.body)
            dusun.name = data['name']
            dusun.code = data['code']
            dusun.description = data.get('description', dusun.description)
            if 'area_size' in data:
                dusun.area_size = data['area_size']
            dusun.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': dusun.id,
                    'name': dusun.name,
                    'code': dusun.code
                }
            })
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def dusun_delete(request, dusun_id):
    """View untuk menghapus dusun"""
    if request.method == 'POST':
        try:
            dusun = get_object_or_404(Dusun, id=dusun_id)
            dusun_name = dusun.name
            dusun.is_active = False
            dusun.save()
            messages.success(request, f'Data dusun {dusun_name} berhasil dihapus.')
            return JsonResponse({'success': True})
        except Exception as e:
            messages.error(request, f'Gagal menghapus dusun: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def dusun_bulk_delete(request):
    """View untuk menghapus multiple dusun sekaligus"""
    if request.method == 'POST':
        try:
            dusun_ids = request.POST.getlist('dusun_ids[]')
            
            if not dusun_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada data yang dipilih'
                }, status=400)
            
            # Get dusun objects
            dusun_list = Dusun.objects.filter(id__in=dusun_ids, is_active=True)
            deleted_count = 0
            errors = []
            
            for dusun in dusun_list:
                try:
                    # Delete related data first (cascade delete)
                    # Delete all families in this dusun
                    Family.objects.filter(dusun=dusun, is_active=True).update(is_active=False)
                    
                    # Delete all residents in this dusun
                    Penduduk.objects.filter(dusun=dusun, is_active=True).update(is_active=False)
                    
                    # Delete all lorongs in this dusun
                    Lorong.objects.filter(dusun=dusun, is_active=True).update(is_active=False)
                    
                    # Soft delete dusun
                    dusun.is_active = False
                    dusun.save()
                    deleted_count += 1
                    
                except Exception as e:
                    errors.append(f'Error menghapus dusun {dusun.name}: {str(e)}')
            
            if errors:
                return JsonResponse({
                    'success': False,
                    'message': 'Beberapa data tidak dapat dihapus',
                    'errors': errors,
                    'deleted_count': deleted_count
                })
            
            return JsonResponse({
                'success': True,
                'deleted_count': deleted_count,
                'message': f'Berhasil menghapus {deleted_count} data dusun'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def dusun_bulk_activate(request):
    """View untuk mengaktifkan multiple dusun sekaligus"""
    if request.method == 'POST':
        try:
            dusun_ids = request.POST.getlist('dusun_ids[]')
            if not dusun_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada data yang dipilih'
                }, status=400)
            
            # Get dusun objects
            dusun_list = Dusun.objects.filter(id__in=dusun_ids)
            activated_count = 0
            
            for dusun in dusun_list:
                if not dusun.is_active:
                    dusun.is_active = True
                    dusun.save()
                    activated_count += 1
            
            return JsonResponse({
                'success': True,
                'activated_count': activated_count,
                'message': f'Berhasil mengaktifkan {activated_count} data dusun'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def dusun_bulk_deactivate(request):
    """View untuk menonaktifkan multiple dusun sekaligus"""
    if request.method == 'POST':
        try:
            dusun_ids = request.POST.getlist('dusun_ids[]')
            if not dusun_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada data yang dipilih'
                }, status=400)
            
            # Get dusun objects
            dusun_list = Dusun.objects.filter(id__in=dusun_ids)
            deactivated_count = 0
            
            for dusun in dusun_list:
                if dusun.is_active:
                    dusun.is_active = False
                    dusun.save()
                    deactivated_count += 1
            
            return JsonResponse({
                'success': True,
                'deactivated_count': deactivated_count,
                'message': f'Berhasil menonaktifkan {deactivated_count} data dusun'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def api_dusun_list(request):
    """API untuk mendapatkan daftar dusun"""
    try:
        dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
        results = []
        for dusun in dusun_list:
            # Count RW and population
            rw_count = dusun.rws.filter(is_active=True).count() if hasattr(dusun, 'rws') else 0
            population_count = dusun.residents.filter(is_active=True, is_alive=True).count()
            
            results.append({
                'id': dusun.id,
                'name': dusun.name,
                'code': dusun.code,
                'description': dusun.description or '',
                'area_size': str(dusun.area_size) if dusun.area_size else '0',
                'kepala_dusun': dusun.kepala_dusun.name if dusun.kepala_dusun else '',
                'jumlah_rw': rw_count,
                'population_count': population_count,
                'is_active': dusun.is_active,
                'alamat_kantor': getattr(dusun, 'alamat_kantor', '') or ''
            })
        
        return JsonResponse({
            'success': True,
            'data': results
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_dusun_detail(request, pk):
    """API untuk mendapatkan detail dusun"""
    try:
        dusun = get_object_or_404(Dusun, id=pk, is_active=True)
        data = {
            'id': dusun.id,
            'name': dusun.name,
            'code': dusun.code,
            'description': dusun.description,
            'area_size': dusun.area_size
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_lorong_list(request):
    """API untuk mendapatkan daftar lorong"""
    try:
        from django.core.paginator import Paginator
        
        lorong_list = Lorong.objects.select_related('dusun').filter(is_active=True).order_by('nama_lorong')
        
        # Filter berdasarkan parameter
        search = request.GET.get('search')
        dusun = request.GET.get('dusun')
        status = request.GET.get('status')
        
        if search:
            lorong_list = lorong_list.filter(
                Q(nama_lorong__icontains=search) |
                Q(kode__icontains=search) |
                Q(ketua_lorong__icontains=search)
            )
        
        if dusun:
            lorong_list = lorong_list.filter(dusun_id=dusun)
            
        if status:
            if status == 'aktif':
                lorong_list = lorong_list.filter(is_active=True)
            elif status == 'tidak_aktif':
                lorong_list = lorong_list.filter(is_active=False)
        
        # Pagination
        page = int(request.GET.get('page', 1))
        per_page = 15
        paginator = Paginator(lorong_list, per_page)
        page_obj = paginator.get_page(page)
        
        # Hitung statistik
        total_lorong = lorong_list.count()
        active_lorong = lorong_list.filter(is_active=True).count()
        total_rt = 0  # RT tidak ada di model Lorong
        total_penduduk = sum(lorong.population_count or 0 for lorong in lorong_list)
        
        results = [{
            'id': lorong.id,
            'nama_lorong': lorong.nama_lorong,
            'kode': lorong.kode,
            'dusun_id': lorong.dusun.id,
            'dusun_name': lorong.dusun.name,
            'ketua_lorong': lorong.ketua_lorong,
            'description': lorong.description or '',
            'length': float(lorong.length) if lorong.length else 0,
            'house_count': lorong.house_count or 0,
            'population_count': lorong.population_count or 0,
            'rt': '-',  # RT tidak ada di model Lorong
            'rw': '-',
            'is_active': lorong.is_active,
            'created_at': lorong.created_at.isoformat() if lorong.created_at else None,
            'updated_at': lorong.updated_at.isoformat() if lorong.updated_at else None
        } for lorong in page_obj]
        
        return JsonResponse({
            'success': True,
            'data': results,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'per_page': per_page,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None
            },
            'statistics': {
                'total_lorong': total_lorong,
                'active_lorong': active_lorong,
                'total_rt': total_rt,
                'total_penduduk': total_penduduk
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_lorong_detail(request, pk):
    """API untuk mendapatkan detail lorong"""
    try:
        lorong = get_object_or_404(Lorong, id=pk)
        data = {
            'id': lorong.id,
            'nama_lorong': lorong.nama_lorong,
            'kode': lorong.kode,
            'dusun': {
                'id': lorong.dusun.id,
                'name': lorong.dusun.name
            },
            'ketua_lorong': lorong.ketua_lorong,
            'description': lorong.description,
            'length': lorong.length,
            'house_count': lorong.house_count,
            'population_count': lorong.population_count,
            'rt': lorong.rt,
            'rw': lorong.rw,
            'is_active': lorong.is_active,
            'created_at': lorong.created_at.isoformat() if lorong.created_at else None,
            'updated_at': lorong.updated_at.isoformat() if lorong.updated_at else None
        }
        return JsonResponse({
            'success': True,
            'data': data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_pelajar_list(request):
    """API untuk mendapatkan daftar pelajar"""
    try:
        pelajar_list = Pelajar.objects.filter(is_active=True)
        
        # Filter berdasarkan parameter
        q = request.GET.get('q')
        dusun = request.GET.get('dusun')
        jenjang = request.GET.get('jenjang')
        
        if q:
            pelajar_list = pelajar_list.filter(
                Q(penduduk__name__icontains=q) |
                Q(penduduk__nik__icontains=q) |
                Q(sekolah__icontains=q) |
                Q(jenjang__icontains=q)
            )
        
        if dusun:
            pelajar_list = pelajar_list.filter(penduduk__dusun_id=dusun)
            
        if jenjang:
            pelajar_list = pelajar_list.filter(jenjang=jenjang)
        
        results = [{
            'id': pelajar.id,
            'name': pelajar.penduduk.nama,
            'nik': pelajar.penduduk.nik,
            'jenjang': pelajar.jenjang,
            'sekolah': pelajar.sekolah,
            'alamat': pelajar.penduduk.alamat,
            'dusun': {
                'id': pelajar.penduduk.dusun.id,
                'name': pelajar.penduduk.dusun.name
            },
            'tahun_masuk': pelajar.tahun_masuk,
            'status': pelajar.status
        } for pelajar in pelajar_list]
        
        return JsonResponse({
            'results': results
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_pelajar_detail(request, pk):
    """API untuk mendapatkan detail pelajar"""
    try:
        pelajar = get_object_or_404(Pelajar, id=pk, is_active=True)
        data = {
            'id': pelajar.id,
            'name': pelajar.penduduk.name,
            'nik': pelajar.penduduk.nik,
            'jenjang': pelajar.jenjang,
            'sekolah': pelajar.sekolah,
            'alamat': pelajar.penduduk.address,
            'dusun': {
                'id': pelajar.penduduk.dusun.id,
                'name': pelajar.penduduk.dusun.name
            },
            'tahun_masuk': pelajar.tahun_masuk,
            'status': pelajar.status
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_disabilitas_detail(request, pk):
    """API untuk mendapatkan detail disabilitas"""
    try:
        disabilitas = get_object_or_404(DisabilitasData, id=pk, is_active=True)
        data = {
            'id': disabilitas.id,
            'penduduk': {
                'id': disabilitas.penduduk.id,
                'name': disabilitas.penduduk.name,
                'nik': disabilitas.penduduk.nik,
                'gender': disabilitas.penduduk.gender,
                'birth_date': disabilitas.penduduk.birth_date.isoformat() if disabilitas.penduduk.birth_date else None,
                'birth_place': disabilitas.penduduk.birth_place,
                'religion': disabilitas.penduduk.religion,
                'education': disabilitas.penduduk.education,
                'occupation': disabilitas.penduduk.occupation,
                'marital_status': disabilitas.penduduk.marital_status,
                'phone_number': disabilitas.penduduk.phone_number,
                'full_address': disabilitas.penduduk.full_address,
                'rt_number': disabilitas.penduduk.rt_number,
                'rw_number': disabilitas.penduduk.rw_number,
                'house_number': disabilitas.penduduk.house_number,
                'kk_number': disabilitas.penduduk.kk_number,
                'relationship_to_head': disabilitas.penduduk.relationship_to_head,
                'dusun': {
                    'id': disabilitas.penduduk.dusun.id,
                    'name': disabilitas.penduduk.dusun.name
                } if disabilitas.penduduk.dusun else None,
                'family': {
                    'id': disabilitas.penduduk.family.id,
                    'head_of_family': disabilitas.penduduk.family.head_of_family.name,
                    'members_count': disabilitas.penduduk.family.members.count()
                } if disabilitas.penduduk.family else None
            },
            'disability_type': {
                'id': disabilitas.disability_type.id,
                'name': disabilitas.disability_type.name
            },
            'severity': disabilitas.severity,
            'description': disabilitas.description,
            'diagnosis_date': disabilitas.diagnosis_date.isoformat() if disabilitas.diagnosis_date else None,
            'needs_assistance': disabilitas.needs_assistance,
            'is_active': disabilitas.is_active,
            'created_at': disabilitas.created_at.isoformat() if disabilitas.created_at else None,
            'updated_at': disabilitas.updated_at.isoformat() if disabilitas.updated_at else None
        }
        return JsonResponse({
            'success': True,
            'data': data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_keluarga_detail(request, pk):
    """API untuk mendapatkan detail keluarga"""
    try:
        keluarga = get_object_or_404(Keluarga, id=pk, is_active=True)
        data = {
            'id': keluarga.id,
            'nomor_kk': keluarga.nomor_kk,
            'nama_kepala_keluarga': keluarga.nama_kepala_keluarga,
            'alamat': keluarga.alamat,
            'dusun': {
                'id': keluarga.dusun.id,
                'name': keluarga.dusun.name
            } if keluarga.dusun else None,
            'rt': keluarga.rt,
            'rw': keluarga.rw,
            'is_active': keluarga.is_active,
            'head_of_family': {
                'id': keluarga.head_of_family.id,
                'name': keluarga.head_of_family.name,
                'nik': keluarga.head_of_family.nik,
                'gender': keluarga.head_of_family.gender,
                'birth_date': keluarga.head_of_family.birth_date.isoformat() if keluarga.head_of_family.birth_date else None,
                'birth_place': keluarga.head_of_family.birth_place,
                'religion': keluarga.head_of_family.religion,
                'education': keluarga.head_of_family.education,
                'occupation': keluarga.head_of_family.occupation,
                'marital_status': keluarga.head_of_family.marital_status,
                'phone_number': keluarga.head_of_family.phone_number,
                'full_address': keluarga.head_of_family.full_address,
                'rt_number': keluarga.head_of_family.rt_number,
                'rw_number': keluarga.head_of_family.rw_number,
                'house_number': keluarga.head_of_family.house_number
            } if keluarga.head_of_family else None,
            'members_count': keluarga.members.count() if hasattr(keluarga, 'members') else 0,
            'created_at': keluarga.created_at.isoformat() if keluarga.created_at else None,
            'updated_at': keluarga.updated_at.isoformat() if keluarga.updated_at else None
        }
        return JsonResponse({
            'success': True,
            'data': data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_population_overview(request):
    """API for population overview statistics"""
    try:
        # Basic population counts
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        male_count = Penduduk.objects.filter(is_active=True, is_alive=True, gender='L').count()
        female_count = Penduduk.objects.filter(is_active=True, is_alive=True, gender='P').count()
        total_families = Family.objects.filter(is_active=True).count()
        
        # Age groups calculation
        today = date.today()
        age_groups = {
            '0-14': 0,
            '15-24': 0,
            '25-54': 0,
            '55-64': 0,
            '65+': 0
        }
        
        residents = Penduduk.objects.filter(is_active=True, is_alive=True)
        for resident in residents:
            age = today.year - resident.birth_date.year
            # Handle leap year edge case for February 29th
            try:
                if resident.birth_date.replace(year=today.year) > today:
                    age -= 1
            except ValueError:
                # Handle February 29th in non-leap year
                if resident.birth_date.month == 2 and resident.birth_date.day == 29:
                    # Use February 28th for comparison
                    birth_date_this_year = resident.birth_date.replace(year=today.year, day=28)
                    if birth_date_this_year > today:
                        age -= 1
                else:
                    # For other cases, just use the original logic
                    if resident.birth_date.replace(year=today.year) > today:
                        age -= 1
                
            if age <= 14:
                age_groups['0-14'] += 1
            elif age <= 24:
                age_groups['15-24'] += 1
            elif age <= 54:
                age_groups['25-54'] += 1
            elif age <= 64:
                age_groups['55-64'] += 1
            else:
                age_groups['65+'] += 1
        
        # Dusun statistics
        dusun_stats = []
        for dusun in Dusun.objects.filter(is_active=True):
            dusun_population = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun
            ).count()
            dusun_male = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='L'
            ).count()
            dusun_female = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='P'
            ).count()
            
            dusun_stats.append({
                'id': dusun.id,
                'name': dusun.name,
                'code': dusun.code,
                'total': dusun_population,
                'male': dusun_male,
                'female': dusun_female,
                'area_size': float(dusun.area_size) if dusun.area_size else 0
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_population': total_population,
                'male_population': male_count,
                'female_population': female_count,
                'total_families': total_families,
                'age_groups': age_groups,
                'dusun_stats': dusun_stats
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@login_required
def pelajar_list(request):
    """View untuk menampilkan daftar pelajar"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    level_filter = request.GET.get('level', '').strip()
    
    # Get all pelajar data
    pelajar_list = Pelajar.objects.select_related(
        'penduduk'
    ).filter(is_active=True).order_by('-created_at')
    
    # Apply search filter
    if search_query:
        pelajar_list = pelajar_list.filter(
            Q(penduduk__name__icontains=search_query) |
            Q(penduduk__nik__icontains=search_query) |
            Q(sekolah__icontains=search_query)
        )
    
    # Apply level filter
    if level_filter:
        pelajar_list = pelajar_list.filter(jenjang=level_filter)
    
    # Pagination
    paginator = Paginator(pelajar_list, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get education levels for filter
    level_options = [
        ('TK', 'Taman Kanak-kanak'),
        ('SD', 'Sekolah Dasar'),
        ('SMP', 'Sekolah Menengah Pertama'),
        ('SMA', 'Sekolah Menengah Atas'),
        ('SMK', 'Sekolah Menengah Kejuruan'),
        ('D1', 'Diploma 1'),
        ('D2', 'Diploma 2'),
        ('D3', 'Diploma 3'),
        ('D4', 'Diploma 4'),
        ('S1', 'Sarjana'),
        ('S2', 'Magister'),
        ('S3', 'Doktor'),
    ]
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'tingkat_filter': level_filter,
        'level_options': level_options,
        'total_count': paginator.count,
        'active_menu': 'references',
        'active_submenu': 'pelajar',
    }
    
    return render(request, 'admin_panel/references/pelajar_list.html', context)

@login_required
def pelajar_detail(request, pelajar_id):
    """View untuk menampilkan detail pelajar"""
    try:
        pelajar = Pelajar.objects.get(id=pelajar_id)
        context = {
            'pelajar': pelajar,
            'page_title': f'Detail Pelajar - {pelajar.penduduk.name}',
        }
        return render(request, 'admin_panel/references/pelajar_detail.html', context)
    except Pelajar.DoesNotExist:
        messages.error(request, 'Data pelajar tidak ditemukan.')
        return redirect('admin_panel:references_pelajar_list')

@login_required
def lorong_list(request):
    """View untuk menampilkan daftar lorong"""
    dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
    context = {
        'dusun_list': dusun_list
    }
    return render(request, 'admin_panel/references/lorong_list.html', context)

@login_required
def lorong_add(request):
    """View untuk menambah data lorong baru"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            dusun = get_object_or_404(Dusun, id=data['dusun_id'])
            
            lorong = Lorong.objects.create(
                dusun=dusun,
                nama_lorong=data['nama_lorong'],
                kode=data['kode'],
                ketua_lorong=data['ketua_lorong'],
                rt_number=data.get('rt_number', ''),
                description=data.get('description', ''),
                length=data.get('length'),
                is_active=data.get('is_active', True)
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Data lorong berhasil ditambahkan',
                'data': {
                    'id': lorong.id,
                    'nama_lorong': lorong.nama_lorong,
                    'dusun': lorong.dusun.name,
                    'kode': lorong.kode,
                    'ketua_lorong': lorong.ketua_lorong
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
    context = {
        'dusun_list': dusun_list
    }
    return render(request, 'admin_panel/references/create/lorong_form.html', context)

@login_required
def lorong_detail(request, lorong_id):
    """View untuk menampilkan detail lorong"""
    try:
        lorong = get_object_or_404(Lorong, id=lorong_id)
        context = {
            'lorong': lorong
        }
        return render(request, 'admin_panel/references/lorong_detail.html', context)
    except Lorong.DoesNotExist:
        messages.error(request, 'Data lorong tidak ditemukan.')
        return redirect('admin_panel:references_lorong_list')

@login_required
def lorong_edit(request, lorong_id):
    """View untuk mengedit lorong"""
    try:
        lorong = get_object_or_404(Lorong, id=lorong_id)
        dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
        context = {
            'lorong': lorong,
            'dusun_list': dusun_list
        }
        return render(request, 'admin_panel/references/create/lorong_form.html', context)
    except Lorong.DoesNotExist:
        messages.error(request, 'Data lorong tidak ditemukan.')
        return redirect('admin_panel:references_lorong_list')

@login_required
def lorong_update(request, lorong_id):
    """View untuk mengupdate data lorong"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            lorong = get_object_or_404(Lorong, id=lorong_id)
            dusun = get_object_or_404(Dusun, id=data['dusun_id'])
            
            lorong.dusun = dusun
            lorong.nama_lorong = data['nama_lorong']
            lorong.kode = data['kode']
            lorong.ketua_lorong = data['ketua_lorong']
            lorong.rt_number = data.get('rt_number', '')
            lorong.description = data.get('description', '')
            lorong.length = data.get('length')
            lorong.is_active = data.get('is_active', True)
            # house_count and population_count will be auto-calculated in model save()
            lorong.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Data lorong berhasil diupdate',
                'data': {
                    'id': lorong.id,
                    'nama_lorong': lorong.nama_lorong,
                    'dusun': lorong.dusun.name,
                    'kode': lorong.kode,
                    'ketua_lorong': lorong.ketua_lorong
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return render(request, 'admin_panel/references/lorong_update.html')

@login_required
def lorong_delete(request, lorong_id):
    """View untuk menghapus data lorong via modal"""
    if request.method == 'POST':
        try:
            lorong = get_object_or_404(Lorong, id=lorong_id)
            lorong_name = lorong.nama_lorong
            
            # Get reason from form data
            reason = request.POST.get('reason', '')
            
            # Soft delete lorong
            lorong.is_active = False
            lorong.save()
            
            # Add success message
            messages.success(request, f'Data lorong {lorong_name} berhasil dihapus')
            
            # Return JSON response for modal
            return JsonResponse({
                'success': True, 
                'message': f'Data lorong {lorong_name} berhasil dihapus'
            })
            
        except Lorong.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Data lorong tidak ditemukan'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    else:
        # For GET requests, redirect to list
        return redirect('admin_panel:references_lorong_list')

@login_required
def lorong_bulk_delete(request):
    """View untuk menghapus multiple lorong sekaligus"""
    if request.method == 'POST':
        try:
            lorong_ids = request.POST.getlist('lorong_ids[]')
            
            if not lorong_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada data yang dipilih'
                }, status=400)
            
            # Get lorong objects
            lorong_list = Lorong.objects.filter(id__in=lorong_ids, is_active=True)
            deleted_count = 0
            errors = []
            
            for lorong in lorong_list:
                try:
                    # Delete related data first (cascade delete)
                    # Delete all residents in this lorong
                    Penduduk.objects.filter(lorong=lorong, is_active=True).update(is_active=False)
                    
                    # Delete all families in this lorong
                    Family.objects.filter(lorong=lorong, is_active=True).update(is_active=False)
                    
                    # Soft delete lorong
                    lorong.is_active = False
                    lorong.save()
                    deleted_count += 1
                    
                except Exception as e:
                    errors.append(f'Error menghapus lorong {lorong.nama_lorong}: {str(e)}')
            
            if errors:
                return JsonResponse({
                    'success': False,
                    'message': 'Beberapa data tidak dapat dihapus',
                    'errors': errors,
                    'deleted_count': deleted_count
                })
            
            return JsonResponse({
                'success': True,
                'deleted_count': deleted_count,
                'message': f'Berhasil menghapus {deleted_count} data lorong'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def lorong_bulk_activate(request):
    """View untuk mengaktifkan multiple lorong sekaligus"""
    if request.method == 'POST':
        try:
            lorong_ids = request.POST.getlist('lorong_ids[]')
            if not lorong_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada data yang dipilih'
                }, status=400)
            
            # Get lorong objects
            lorong_list = Lorong.objects.filter(id__in=lorong_ids)
            activated_count = 0
            
            for lorong in lorong_list:
                if not lorong.is_active:
                    lorong.is_active = True
                    lorong.save()
                    activated_count += 1
            
            # Success message will be handled by JavaScript
            
            return JsonResponse({
                'success': True,
                'activated_count': activated_count,
                'message': f'Berhasil mengaktifkan {activated_count} data lorong'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def lorong_bulk_deactivate(request):
    """View untuk menonaktifkan multiple lorong sekaligus"""
    if request.method == 'POST':
        try:
            lorong_ids = request.POST.getlist('lorong_ids[]')
            if not lorong_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada data yang dipilih'
                }, status=400)
            
            # Get lorong objects
            lorong_list = Lorong.objects.filter(id__in=lorong_ids)
            deactivated_count = 0
            
            for lorong in lorong_list:
                if lorong.is_active:
                    lorong.is_active = False
                    lorong.save()
                    deactivated_count += 1
            
            # Success message will be handled by JavaScript
            
            return JsonResponse({
                'success': True,
                'deactivated_count': deactivated_count,
                'message': f'Berhasil menonaktifkan {deactivated_count} data lorong'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

# Keluarga Views
@login_required
def keluarga_list(request):
    """View untuk menampilkan daftar keluarga"""
    keluarga_list = Keluarga.objects.select_related('dusun').all()
    dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
    context = {
        'keluarga_list': keluarga_list,
        'dusun_list': dusun_list
    }
    return render(request, 'admin_panel/references/keluarga_list.html', context)

@login_required
def keluarga_add(request):
    """View untuk menambah data keluarga baru"""
    if request.method == 'POST':
        try:
            dusun = get_object_or_404(Dusun, id=request.POST.get('dusun'))
            
            keluarga = Keluarga.objects.create(
                dusun=dusun,
                nomor_kk=request.POST.get('nomor_kk'),
                nama_kepala_keluarga=request.POST.get('nama_kepala_keluarga'),
                alamat=request.POST.get('alamat'),
                rt=request.POST.get('rt'),
                rw=request.POST.get('rw'),
                is_active=request.POST.get('is_active') == 'on'
            )
            
            # Handle anggota keluarga
            anggota_ids = request.POST.getlist('anggota_keluarga')
            if anggota_ids:
                Penduduk.objects.filter(id__in=anggota_ids).update(kk_number=keluarga.nomor_kk)
            
            messages.success(request, 'Data keluarga berhasil ditambahkan')
            return redirect('admin_panel:references_keluarga_detail', keluarga_id=keluarga.id)
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
    penduduk_list = Penduduk.objects.filter(kk_number__isnull=True).order_by('name')
    context = {
        'dusun_list': dusun_list,
        'penduduk_list': penduduk_list,
        'form_title': 'Tambah Keluarga',
        'form_subtitle': 'Lengkapi data keluarga baru'
    }
    return render(request, 'admin_panel/references/create/keluarga_form.html', context)

@login_required
def keluarga_detail(request, keluarga_id):
    """View untuk menampilkan detail keluarga"""
    try:
        keluarga = get_object_or_404(Keluarga, id=keluarga_id)
        context = {
            'keluarga': keluarga
        }
        return render(request, 'admin_panel/references/keluarga_detail.html', context)
    except Keluarga.DoesNotExist:
        messages.error(request, 'Data keluarga tidak ditemukan.')
        return redirect('admin_panel:references_keluarga_list')

@login_required
def keluarga_edit(request, keluarga_id):
    """View untuk mengedit keluarga (hanya anggota)"""
    try:
        keluarga = get_object_or_404(Keluarga, id=keluarga_id)
        dusun_list = Dusun.objects.filter(is_active=True).order_by('name')
        
        # Get all penduduk (both with and without kk_number)
        penduduk_list = Penduduk.objects.all().order_by('name')
        
        # Get current anggota keluarga
        current_anggota = Penduduk.objects.filter(kk_number=keluarga.nomor_kk)
        
        # Get kepala keluarga ID for edit link
        kepala_keluarga = Penduduk.objects.filter(
            kk_number=keluarga.nomor_kk,
            relationship_to_head='KEPALA_KELUARGA'
        ).first()
        
        context = {
            'keluarga': keluarga,
            'dusun_list': dusun_list,
            'penduduk_list': penduduk_list,
            'current_anggota': current_anggota,
            'kepala_keluarga_id': kepala_keluarga.id if kepala_keluarga else None,
            'form_title': 'Kelola Anggota Keluarga',
            'form_subtitle': f'Kelola anggota keluarga {keluarga.nama_kepala_keluarga}'
        }
        return render(request, 'admin_panel/references/create/keluarga_form.html', context)
    except Keluarga.DoesNotExist:
        messages.error(request, 'Data keluarga tidak ditemukan.')
        return redirect('admin_panel:references_keluarga_list')

@login_required
def keluarga_update(request, keluarga_id):
    """View untuk mengupdate data keluarga (deprecated - redirect to anggota update)"""
    return redirect('admin_panel:references_keluarga_update_anggota', keluarga_id=keluarga_id)

@login_required
def keluarga_update_anggota(request, keluarga_id):
    """View untuk mengupdate anggota keluarga saja (data keluarga readonly)"""
    try:
        keluarga = get_object_or_404(Keluarga, id=keluarga_id)
        if request.method == 'POST':
            # Handle anggota keluarga
            # First, remove all current anggota
            Penduduk.objects.filter(kk_number=keluarga.nomor_kk).update(kk_number=None)
            
            # Add new anggota
            anggota_ids = request.POST.getlist('anggota_keluarga')
            if anggota_ids:
                Penduduk.objects.filter(id__in=anggota_ids).update(kk_number=keluarga.nomor_kk)
            
            messages.success(request, 'Anggota keluarga berhasil diupdate')
            return redirect('admin_panel:references_keluarga_detail', keluarga_id=keluarga.id)
        return redirect('admin_panel:references_keluarga_list')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('admin_panel:references_keluarga_list')

@login_required
def keluarga_delete(request, keluarga_id):
    """View untuk menghapus data keluarga via modal"""
    if request.method == 'POST':
        try:
            keluarga = get_object_or_404(Keluarga, id=keluarga_id)
            keluarga_name = keluarga.nama_kepala_keluarga
            
            # Get reason from form data
            reason = request.POST.get('reason', '')
            
            # Soft delete keluarga
            keluarga.is_active = False
            keluarga.save()
            
            # Return JSON response for modal
            return JsonResponse({
                'success': True, 
                'message': f'Data keluarga {keluarga_name} berhasil dihapus'
            })
            
        except Keluarga.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Data keluarga tidak ditemukan'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    else:
        # For GET requests, redirect to list
        return redirect('admin_panel:references_keluarga_list')

@login_required
def disabilitas_list(request):
    """View untuk menampilkan daftar disabilitas"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    type_filter = request.GET.get('type', '').strip()
    
    # Get all disabilitas data
    disabilitas_list = DisabilitasData.objects.select_related(
        'penduduk', 'disability_type'
    ).filter(is_active=True).order_by('-created_at')
    
    # Apply search filter
    if search_query:
        disabilitas_list = disabilitas_list.filter(
            Q(penduduk__name__icontains=search_query) |
            Q(penduduk__nik__icontains=search_query)
        )
    
    # Apply type filter
    if type_filter:
        disabilitas_list = disabilitas_list.filter(disability_type_id=type_filter)
    
    # Pagination
    paginator = Paginator(disabilitas_list, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get disability types for filter
    type_options = DisabilitasType.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'type_filter': type_filter,
        'type_options': type_options,
        'total_count': paginator.count,
        'active_menu': 'references',
        'active_submenu': 'disabilitas',
    }
    
    return render(request, 'admin_panel/references/disabilitas_list.html', context)

@login_required
def disabilitas_add(request):
    """View untuk menambah data disabilitas baru"""
    if request.method == 'POST':
        try:
            # Get form data
            penduduk_id = request.POST.get('penduduk_id')
            disability_type_id = request.POST.get('disability_type_id')
            severity = request.POST.get('severity', 'RINGAN')
            description = request.POST.get('description', '')
            diagnosis_date = request.POST.get('diagnosis_date')
            needs_assistance = request.POST.get('needs_assistance') == 'on'
            is_active = request.POST.get('is_active') == 'on' if 'is_active' in request.POST else True
            
            # Validate required fields
            if not penduduk_id:
                messages.error(request, 'Penduduk harus dipilih.')
                return redirect('admin_panel:references_disabilitas_add')
            
            if not disability_type_id:
                messages.error(request, 'Jenis disabilitas harus dipilih.')
                return redirect('admin_panel:references_disabilitas_add')
            
            # Get objects
            penduduk = get_object_or_404(Penduduk, id=penduduk_id)
            disability_type = get_object_or_404(DisabilitasType, id=disability_type_id)
            
            # Check if penduduk already has disabilitas data
            if DisabilitasData.objects.filter(penduduk=penduduk, is_active=True).exists():
                messages.warning(request, f'Penduduk {penduduk.name} sudah memiliki data disabilitas aktif.')
                return redirect('admin_panel:references_disabilitas_add')
            
            # Create disabilitas
            disabilitas = DisabilitasData.objects.create(
                penduduk=penduduk,
                disability_type=disability_type,
                severity=severity,
                description=description,
                diagnosis_date=diagnosis_date if diagnosis_date else None,
                needs_assistance=needs_assistance,
                is_active=is_active
            )
            
            messages.success(request, f'Data disabilitas untuk {penduduk.name} berhasil ditambahkan.')
            return redirect('admin_panel:references_disabilitas_list')
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_panel:references_disabilitas_add')
    
    # GET request - show form
    type_options = DisabilitasType.objects.filter(is_active=True).order_by('name')
    context = {
        'type_options': type_options,
        'form_title': 'Tambah Disabilitas',
        'form_subtitle': 'Lengkapi data disabilitas baru',
        'form_icon': 'plus'
    }
    return render(request, 'admin_panel/references/create/disabilitas_form.html', context)

@login_required
def disabilitas_detail(request, disabilitas_id):
    """View untuk menampilkan detail disabilitas"""
    try:
        disabilitas = get_object_or_404(DisabilitasData, id=disabilitas_id)
        context = {
            'disabilitas': disabilitas
        }
        return render(request, 'admin_panel/references/disabilitas_detail.html', context)
    except DisabilitasData.DoesNotExist:
        messages.error(request, 'Data disabilitas tidak ditemukan.')
        return redirect('admin_panel:references_disabilitas_list')

@login_required
def disabilitas_edit(request, disabilitas_id):
    """View untuk mengedit disabilitas"""
    try:
        disabilitas = get_object_or_404(DisabilitasData, id=disabilitas_id)
        context = {
            'disabilitas': disabilitas
        }
        return render(request, 'admin_panel/references/create/disabilitas_form.html', context)
    except DisabilitasData.DoesNotExist:
        messages.error(request, 'Data disabilitas tidak ditemukan.')
        return redirect('admin_panel:references_disabilitas_list')

@login_required
def disabilitas_update(request, disabilitas_id):
    """View untuk mengupdate data disabilitas"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            disabilitas = get_object_or_404(DisabilitasData, id=disabilitas_id)
            penduduk = get_object_or_404(Penduduk, id=data['penduduk_id'])
            disability_type = get_object_or_404(DisabilitasType, id=data['disability_type_id'])
            
            disabilitas.penduduk = penduduk
            disabilitas.disability_type = disability_type
            disabilitas.severity = data['severity']
            disabilitas.description = data.get('description', '')
            disabilitas.diagnosis_date = data.get('diagnosis_date')
            disabilitas.needs_assistance = data.get('needs_assistance', False)
            disabilitas.is_active = data.get('is_active', True)
            disabilitas.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Data disabilitas berhasil diupdate',
                'data': {
                    'id': disabilitas.id,
                    'penduduk': disabilitas.penduduk.name,
                    'disability_type': disabilitas.disability_type.name,
                    'severity': disabilitas.get_severity_display()
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return render(request, 'admin_panel/references/disabilitas_update.html')

@login_required
def disabilitas_delete(request, disabilitas_id):
    """View untuk menghapus data disabilitas"""
    if request.method == 'POST':
        try:
            disabilitas = get_object_or_404(DisabilitasData, id=disabilitas_id)
            disabilitas.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Data disabilitas berhasil dihapus'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return JsonResponse({
        'success': False,
        'message': 'Method not allowed'
    }, status=405)

@login_required
def pelajar_add(request):
    """View untuk menambah data pelajar baru"""
    if request.method == 'POST':
        try:
            penduduk_id = request.POST.get('penduduk_id')
            if not penduduk_id:
                messages.error(request, 'Penduduk harus dipilih.')
                return render(request, 'admin_panel/references/create/pelajar_form.html')
            
            # Cek apakah penduduk sudah terdaftar sebagai pelajar
            existing_pelajar = Pelajar.objects.filter(penduduk_id=penduduk_id, is_active=True).first()
            if existing_pelajar:
                messages.warning(request, f'Penduduk {existing_pelajar.penduduk.name} sudah terdaftar sebagai pelajar dengan jenjang {existing_pelajar.get_jenjang_display()}.')
                return render(request, 'admin_panel/references/create/pelajar_form.html')
            
            penduduk = get_object_or_404(Penduduk, id=penduduk_id)
            pelajar = Pelajar.objects.create(
                penduduk=penduduk,
                jenjang=request.POST.get('jenjang'),
                sekolah=request.POST.get('sekolah'),
                tahun_masuk=request.POST.get('tahun_masuk') or None,
                status=request.POST.get('status', 'aktif'),
                keterangan=request.POST.get('keterangan', ''),
                is_active=request.POST.get('is_active') == 'on',
                created_by=request.user
            )
            messages.success(request, f'Data pelajar {pelajar.penduduk.name} berhasil ditambahkan.')
            return redirect('admin_panel:references_pelajar_list')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return render(request, 'admin_panel/references/create/pelajar_form.html')
    
    context = {
        'form_title': 'Tambah Pelajar',
        'form_subtitle': 'Lengkapi data pelajar baru',
        'form_icon': 'plus'
    }
    return render(request, 'admin_panel/references/create/pelajar_form.html', context)

@login_required
def pelajar_edit(request, pelajar_id):
    """View untuk mengedit pelajar"""
    try:
        pelajar = get_object_or_404(Pelajar, id=pelajar_id)
        context = {
            'pelajar': pelajar,
            'form_title': 'Edit Pelajar',
            'form_subtitle': f'Edit data pelajar {pelajar.penduduk.name}',
            'form_icon': 'edit'
        }
        return render(request, 'admin_panel/references/create/pelajar_form.html', context)
    except Pelajar.DoesNotExist:
        messages.error(request, 'Data pelajar tidak ditemukan.')
        return redirect('admin_panel:references_pelajar_list')

@login_required
def pelajar_update(request, pelajar_id):
    """View untuk mengupdate data pelajar"""
    try:
        pelajar = get_object_or_404(Pelajar, id=pelajar_id)
        if request.method == 'POST':
            penduduk_id = request.POST.get('penduduk_id')
            if penduduk_id and penduduk_id != str(pelajar.penduduk.id):
                # Cek apakah penduduk baru sudah terdaftar sebagai pelajar
                existing_pelajar = Pelajar.objects.filter(penduduk_id=penduduk_id, is_active=True).exclude(id=pelajar_id).first()
                if existing_pelajar:
                    messages.warning(request, f'Penduduk {existing_pelajar.penduduk.name} sudah terdaftar sebagai pelajar dengan jenjang {existing_pelajar.get_jenjang_display()}.')
                    return redirect('admin_panel:references_pelajar_edit', pelajar_id=pelajar_id)
                pelajar.penduduk = get_object_or_404(Penduduk, id=penduduk_id)
            
            pelajar.jenjang = request.POST.get('jenjang', pelajar.jenjang)
            pelajar.sekolah = request.POST.get('sekolah', pelajar.sekolah)
            pelajar.tahun_masuk = request.POST.get('tahun_masuk') or None
            pelajar.status = request.POST.get('status', pelajar.status)
            pelajar.keterangan = request.POST.get('keterangan', pelajar.keterangan)
            pelajar.is_active = request.POST.get('is_active') == 'on'
            pelajar.updated_by = request.user
            pelajar.save()
            
            messages.success(request, f'Data pelajar {pelajar.penduduk.name} berhasil diupdate.')
            return redirect('admin_panel:references_pelajar_list')
        return redirect('admin_panel:references_pelajar_list')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('admin_panel:references_pelajar_edit', pelajar_id=pelajar_id)

@login_required
def pelajar_delete(request, pelajar_id):
    """View untuk menghapus data pelajar"""
    if request.method == 'POST':
        try:
            pelajar = get_object_or_404(Pelajar, id=pelajar_id)
            pelajar.is_active = False
            pelajar.updated_by = request.user
            pelajar.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def api_pelajar_list(request):
    """API untuk mendapatkan daftar pelajar"""
    try:
        pelajar_list = Pelajar.objects.filter(is_active=True).select_related('penduduk').values(
            'id', 'penduduk__name', 'jenjang', 'sekolah', 'status'
        )
        return JsonResponse({
            'success': True,
            'data': list(pelajar_list)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def pelajar_create(request):
    """View untuk membuat data pelajar baru"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            penduduk = get_object_or_404(Penduduk, id=data['penduduk_id'])
            pelajar = Pelajar.objects.create(
                penduduk=penduduk,
                jenjang=data['jenjang'],
                sekolah=data['sekolah'],
                tahun_masuk=data['tahun_masuk'],
                status=data.get('status', 'aktif'),
                keterangan=data.get('keterangan', ''),
                is_active=True,
                created_by=request.user
            )
            return JsonResponse({
                'success': True,
                'data': {
                    'id': pelajar.id,
                    'penduduk': pelajar.penduduk.name,
                    'jenjang': pelajar.get_jenjang_display(),
                    'sekolah': pelajar.sekolah
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

# Export functions for pelajar
@login_required
def pelajar_export_excel(request):
    """Export pelajar data to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        # Get pelajar data with penduduk information
        pelajar_list = Pelajar.objects.select_related('penduduk', 'penduduk__dusun').filter(is_active=True).order_by('penduduk__name')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pelajar"
        
        # Headers
        headers = [
            'No', 'Nama Lengkap', 'NIK', 'Jenis Kelamin', 'Usia', 'Tempat Lahir', 'Tanggal Lahir',
            'Agama', 'Pendidikan', 'Pekerjaan', 'Status Perkawinan', 'No. Telepon',
            'Dusun', 'RT', 'RW', 'Alamat Lengkap', 'No. KK', 'Hubungan dalam Keluarga',
            'Jenjang Pendidikan', 'Nama Sekolah', 'Tahun Masuk', 'Status Pelajar', 'Keterangan'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, pelajar in enumerate(pelajar_list, 2):
            ws.cell(row=row, column=1, value=row-1)  # No
            ws.cell(row=row, column=2, value=pelajar.penduduk.name)
            ws.cell(row=row, column=3, value=pelajar.penduduk.nik)
            ws.cell(row=row, column=4, value=pelajar.penduduk.get_gender_display())
            ws.cell(row=row, column=5, value=pelajar.penduduk.age)
            ws.cell(row=row, column=6, value=pelajar.penduduk.birth_place or '')
            ws.cell(row=row, column=7, value=pelajar.penduduk.birth_date.strftime('%d/%m/%Y') if pelajar.penduduk.birth_date else '')
            ws.cell(row=row, column=8, value=pelajar.penduduk.religion or '')
            ws.cell(row=row, column=9, value=pelajar.penduduk.get_education_display() if pelajar.penduduk.education else '')
            ws.cell(row=row, column=10, value=pelajar.penduduk.occupation or '')
            ws.cell(row=row, column=11, value=pelajar.penduduk.get_marital_status_display() if pelajar.penduduk.marital_status else '')
            ws.cell(row=row, column=12, value=pelajar.penduduk.phone_number or '')
            ws.cell(row=row, column=13, value=pelajar.penduduk.dusun.name if pelajar.penduduk.dusun else '')
            ws.cell(row=row, column=14, value=pelajar.penduduk.rt_number or '')
            ws.cell(row=row, column=15, value=pelajar.penduduk.rw_number or '')
            ws.cell(row=row, column=16, value=pelajar.penduduk.full_address or '')
            ws.cell(row=row, column=17, value=pelajar.penduduk.kk_number or '')
            ws.cell(row=row, column=18, value=pelajar.penduduk.get_relationship_to_head_display() if pelajar.penduduk.relationship_to_head else '')
            ws.cell(row=row, column=19, value=pelajar.get_jenjang_display())
            ws.cell(row=row, column=20, value=pelajar.sekolah)
            ws.cell(row=row, column=21, value=pelajar.tahun_masuk or '')
            ws.cell(row=row, column=22, value=pelajar.get_status_display())
            ws.cell(row=row, column=23, value=pelajar.keterangan or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="data_pelajar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting Excel: {str(e)}')
        return redirect('admin_panel:references_pelajar_list')

@login_required
def pelajar_export_csv(request):
    """Export pelajar data to CSV"""
    try:
        import csv
        from django.http import HttpResponse
        
        # Get pelajar data
        pelajar_list = Pelajar.objects.select_related('penduduk', 'penduduk__dusun').filter(is_active=True).order_by('penduduk__name')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="data_pelajar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for Excel compatibility
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'No', 'Nama Lengkap', 'NIK', 'Jenis Kelamin', 'Usia', 'Tempat Lahir', 'Tanggal Lahir',
            'Agama', 'Pendidikan', 'Pekerjaan', 'Status Perkawinan', 'No. Telepon',
            'Dusun', 'RT', 'RW', 'Alamat Lengkap', 'No. KK', 'Hubungan dalam Keluarga',
            'Jenjang Pendidikan', 'Nama Sekolah', 'Tahun Masuk', 'Status Pelajar', 'Keterangan'
        ])
        
        # Data rows
        for row, pelajar in enumerate(pelajar_list, 1):
            writer.writerow([
                row,
                pelajar.penduduk.name,
                pelajar.penduduk.nik,
                pelajar.penduduk.get_gender_display(),
                pelajar.penduduk.age,
                pelajar.penduduk.birth_place or '',
                pelajar.penduduk.birth_date.strftime('%d/%m/%Y') if pelajar.penduduk.birth_date else '',
                pelajar.penduduk.religion or '',
                pelajar.penduduk.get_education_display() if pelajar.penduduk.education else '',
                pelajar.penduduk.occupation or '',
                pelajar.penduduk.get_marital_status_display() if pelajar.penduduk.marital_status else '',
                pelajar.penduduk.phone_number or '',
                pelajar.penduduk.dusun.name if pelajar.penduduk.dusun else '',
                pelajar.penduduk.rt_number or '',
                pelajar.penduduk.rw_number or '',
                pelajar.penduduk.full_address or '',
                pelajar.penduduk.kk_number or '',
                pelajar.penduduk.get_relationship_to_head_display() if pelajar.penduduk.relationship_to_head else '',
                pelajar.get_jenjang_display(),
                pelajar.sekolah,
                pelajar.tahun_masuk or '',
                pelajar.get_status_display(),
                pelajar.keterangan or ''
            ])
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting CSV: {str(e)}')
        return redirect('admin_panel:references_pelajar_list')

@login_required
def pelajar_export_json(request):
    """Export pelajar data to JSON"""
    try:
        from django.http import JsonResponse
        
        # Get pelajar data
        pelajar_list = Pelajar.objects.select_related('penduduk', 'penduduk__dusun').filter(is_active=True).order_by('penduduk__name')
        
        data = []
        for pelajar in pelajar_list:
            data.append({
                'nama_lengkap': pelajar.penduduk.name,
                'nik': pelajar.penduduk.nik,
                'jenis_kelamin': pelajar.penduduk.get_gender_display(),
                'usia': pelajar.penduduk.age,
                'tempat_lahir': pelajar.penduduk.birth_place or '',
                'tanggal_lahir': pelajar.penduduk.birth_date.strftime('%d/%m/%Y') if pelajar.penduduk.birth_date else '',
                'agama': pelajar.penduduk.religion or '',
                'pendidikan': pelajar.penduduk.get_education_display() if pelajar.penduduk.education else '',
                'pekerjaan': pelajar.penduduk.occupation or '',
                'status_perkawinan': pelajar.penduduk.get_marital_status_display() if pelajar.penduduk.marital_status else '',
                'no_telepon': pelajar.penduduk.phone_number or '',
                'dusun': pelajar.penduduk.dusun.name if pelajar.penduduk.dusun else '',
                'rt': pelajar.penduduk.rt_number or '',
                'rw': pelajar.penduduk.rw_number or '',
                'alamat_lengkap': pelajar.penduduk.full_address or '',
                'no_kk': pelajar.penduduk.kk_number or '',
                'hubungan_dalam_keluarga': pelajar.penduduk.get_relationship_to_head_display() if pelajar.penduduk.relationship_to_head else '',
                'jenjang_pendidikan': pelajar.get_jenjang_display(),
                'nama_sekolah': pelajar.sekolah,
                'tahun_masuk': pelajar.tahun_masuk,
                'status_pelajar': pelajar.get_status_display(),
                'keterangan': pelajar.keterangan or ''
            })
        
        response = JsonResponse({
            'data': data,
            'total': len(data),
            'exported_at': datetime.now().isoformat()
        }, json_dumps_params={'indent': 2, 'ensure_ascii': False})
        
        response['Content-Disposition'] = f'attachment; filename="data_pelajar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting JSON: {str(e)}')
        return redirect('admin_panel:references_pelajar_list')

@login_required
def pelajar_export_pdf(request):
    """Export pelajar data to PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        
        # Get pelajar data
        pelajar_list = Pelajar.objects.select_related('penduduk', 'penduduk__dusun').filter(is_active=True).order_by('penduduk__name')
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="data_pelajar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        story.append(Paragraph("Data Pelajar Desa", title_style))
        story.append(Spacer(1, 12))
        
        # Table data
        table_data = [['No', 'Nama', 'NIK', 'Jenjang', 'Sekolah', 'Status']]
        
        for i, pelajar in enumerate(pelajar_list, 1):
            table_data.append([
                str(i),
                pelajar.penduduk.name[:20] + '...' if len(pelajar.penduduk.name) > 20 else pelajar.penduduk.name,
                pelajar.penduduk.nik,
                pelajar.get_jenjang_display(),
                pelajar.sekolah[:15] + '...' if len(pelajar.sekolah) > 15 else pelajar.sekolah,
                pelajar.get_status_display()
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting PDF: {str(e)}')
        return redirect('admin_panel:references_pelajar_list')

# Statistics functions for pelajar
@login_required
def pelajar_statistics(request):
    """Statistics dashboard for pelajar"""
    try:
        from django.db.models import Count, Q
        from collections import defaultdict
        
        # Get basic statistics
        total_pelajar = Pelajar.objects.filter(is_active=True).count()
        
        # Statistics by jenjang
        jenjang_stats = Pelajar.objects.filter(is_active=True).values('jenjang').annotate(
            count=Count('id')
        ).order_by('jenjang')
        
        # Statistics by status
        status_stats = Pelajar.objects.filter(is_active=True).values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        # Statistics by gender
        gender_stats = Pelajar.objects.filter(is_active=True).select_related('penduduk').values(
            'penduduk__gender'
        ).annotate(count=Count('id')).order_by('penduduk__gender')
        
        # Statistics by dusun
        dusun_stats = Pelajar.objects.filter(is_active=True).select_related('penduduk__dusun').values(
            'penduduk__dusun__name'
        ).annotate(count=Count('id')).order_by('penduduk__dusun__name')
        
        # Age distribution
        age_stats = []
        for pelajar in Pelajar.objects.filter(is_active=True).select_related('penduduk'):
            age = pelajar.penduduk.age
            if age:
                age_group = f"{(age//5)*5}-{((age//5)*5)+4}"
                age_stats.append(age_group)
        
        age_distribution = defaultdict(int)
        for age_group in age_stats:
            age_distribution[age_group] += 1
        
        # School statistics
        school_stats = Pelajar.objects.filter(is_active=True).values('sekolah').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        context = {
            'total_pelajar': total_pelajar,
            'jenjang_stats': jenjang_stats,
            'status_stats': status_stats,
            'gender_stats': gender_stats,
            'dusun_stats': dusun_stats,
            'age_distribution': dict(age_distribution),
            'school_stats': school_stats,
            'active_menu': 'references',
            'active_submenu': 'pelajar',
        }
        
        return render(request, 'admin_panel/references/pelajar_statistics.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading statistics: {str(e)}')
        return redirect('admin_panel:references_pelajar_list')

@login_required
def api_pelajar_statistics(request):
    """API endpoint for pelajar statistics"""
    try:
        from django.db.models import Count
        from collections import defaultdict
        
        # Get basic statistics
        total_pelajar = Pelajar.objects.filter(is_active=True).count()
        
        # Statistics by jenjang
        jenjang_data = []
        for jenjang in Pelajar.JENJANG_CHOICES:
            count = Pelajar.objects.filter(is_active=True, jenjang=jenjang[0]).count()
            jenjang_data.append({
                'jenjang': jenjang[1],
                'count': count,
                'percentage': round((count / total_pelajar * 100), 2) if total_pelajar > 0 else 0
            })
        
        # Statistics by status
        status_data = []
        for status in Pelajar.STATUS_CHOICES:
            count = Pelajar.objects.filter(is_active=True, status=status[0]).count()
            status_data.append({
                'status': status[1],
                'count': count,
                'percentage': round((count / total_pelajar * 100), 2) if total_pelajar > 0 else 0
            })
        
        # Statistics by gender
        gender_data = []
        for gender in [('L', 'Laki-laki'), ('P', 'Perempuan')]:
            count = Pelajar.objects.filter(is_active=True, penduduk__gender=gender[0]).count()
            gender_data.append({
                'gender': gender[1],
                'count': count,
                'percentage': round((count / total_pelajar * 100), 2) if total_pelajar > 0 else 0
            })
        
        # Age distribution
        age_stats = []
        for pelajar in Pelajar.objects.filter(is_active=True).select_related('penduduk'):
            age = pelajar.penduduk.age
            if age:
                age_group = f"{(age//5)*5}-{((age//5)*5)+4}"
                age_stats.append(age_group)
        
        age_distribution = defaultdict(int)
        for age_group in age_stats:
            age_distribution[age_group] += 1
        
        age_data = []
        for age_group, count in sorted(age_distribution.items()):
            age_data.append({
                'age_group': age_group,
                'count': count,
                'percentage': round((count / total_pelajar * 100), 2) if total_pelajar > 0 else 0
            })
        
        # School statistics
        school_stats = Pelajar.objects.filter(is_active=True).values('sekolah').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        school_data = []
        for school in school_stats:
            school_data.append({
                'sekolah': school['sekolah'],
                'count': school['count'],
                'percentage': round((school['count'] / total_pelajar * 100), 2) if total_pelajar > 0 else 0
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_pelajar': total_pelajar,
                'jenjang_distribution': jenjang_data,
                'status_distribution': status_data,
                'gender_distribution': gender_data,
                'age_distribution': age_data,
                'school_distribution': school_data
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def api_population_demographics(request):
    """API for detailed demographics data"""
    try:
        # Age distribution with percentages
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        
        if total_population == 0:
            return JsonResponse({
                'success': True,
                'data': {
                    'age_distribution': {},
                    'marital_status': {},
                    'religion': {},
                    'disability': {},
                    'blood_type': {}
                }
            })
        
        # Age distribution
        today = date.today()
        age_distribution = {
            '0-14': 0,
            '15-24': 0,
            '25-54': 0,
            '55-64': 0,
            '65+': 0
        }
        
        residents = Penduduk.objects.filter(is_active=True, is_alive=True)
        for resident in residents:
            age = today.year - resident.birth_date.year
            # Handle leap year edge case for February 29th
            try:
                if resident.birth_date.replace(year=today.year) > today:
                    age -= 1
            except ValueError:
                # Handle February 29th in non-leap year
                if resident.birth_date.month == 2 and resident.birth_date.day == 29:
                    # Use February 28th for comparison
                    birth_date_this_year = resident.birth_date.replace(year=today.year, day=28)
                    if birth_date_this_year > today:
                        age -= 1
                else:
                    # For other cases, just use the original logic
                    if resident.birth_date.replace(year=today.year) > today:
                        age -= 1
                
            if age <= 14:
                age_distribution['0-14'] += 1
            elif age <= 24:
                age_distribution['15-24'] += 1
            elif age <= 54:
                age_distribution['25-54'] += 1
            elif age <= 64:
                age_distribution['55-64'] += 1
            else:
                age_distribution['65+'] += 1
        
        # Convert to percentages
        for key in age_distribution:
            count = age_distribution[key]
            percentage = (count / total_population * 100) if total_population > 0 else 0
            age_distribution[key] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        # Marital status distribution
        marital_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True
        ).values('marital_status').annotate(count=Count('id'))
        
        marital_status = {}
        for stat in marital_stats:
            status = stat['marital_status']
            count = stat['count']
            percentage = (count / total_population * 100) if total_population > 0 else 0
            marital_status[status] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        # Religion distribution
        religion_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True
        ).values('religion').annotate(count=Count('id'))
        
        religion = {}
        for stat in religion_stats:
            rel = stat['religion']
            count = stat['count']
            percentage = (count / total_population * 100) if total_population > 0 else 0
            religion[rel] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        # Disability statistics
        disability_count = DisabilitasData.objects.filter(
            is_active=True, penduduk__is_active=True, penduduk__is_alive=True
        ).values('penduduk').distinct().count()
        
        disability = {
            'with_disability': {
                'count': disability_count,
                'percentage': round((disability_count / total_population * 100), 1) if total_population > 0 else 0
            },
            'without_disability': {
                'count': total_population - disability_count,
                'percentage': round(((total_population - disability_count) / total_population * 100), 1) if total_population > 0 else 0
            }
        }
        
        # Blood type distribution
        blood_type_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True, blood_type__isnull=False
        ).exclude(blood_type='').values('blood_type').annotate(count=Count('id'))
        
        blood_type = {}
        total_with_blood_type = sum([stat['count'] for stat in blood_type_stats])
        
        for stat in blood_type_stats:
            bt = stat['blood_type']
            count = stat['count']
            percentage = (count / total_with_blood_type * 100) if total_with_blood_type > 0 else 0
            blood_type[bt] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        return JsonResponse({
            'success': True,
            'data': {
                'age_distribution': age_distribution,
                'marital_status': marital_status,
                'religion': religion,
                'disability': disability,
                'blood_type': blood_type
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
def api_population_education(request):
    """API for education statistics"""
    try:
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        
        # Education distribution
        education_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True, education__isnull=False
        ).exclude(education='').values('education').annotate(count=Count('id'))
        
        education = {}
        for stat in education_stats:
            edu = stat['education']
            count = stat['count']
            percentage = (count / total_population * 100) if total_population > 0 else 0
            education[edu] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        return JsonResponse({
            'success': True,
            'data': {
                'education': education
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
def api_population_occupation(request):
    """API for occupation statistics"""
    try:
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        
        # Occupation distribution
        occupation_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True, occupation__isnull=False
        ).exclude(occupation='').values('occupation').annotate(count=Count('id'))
        
        occupation = {}
        for stat in occupation_stats:
            occ = stat['occupation']
            count = stat['count']
            percentage = (count / total_population * 100) if total_population > 0 else 0
            occupation[occ] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        return JsonResponse({
            'success': True,
            'data': {
                'occupation': occupation
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
def api_dusun_details(request):
    """API for detailed dusun information"""
    try:
        dusun_data = []
        
        for dusun in Dusun.objects.filter(is_active=True):
            # Population by gender
            male_count = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='L'
            ).count()
            female_count = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='P'
            ).count()
            total_count = male_count + female_count
            
            # Age groups in this dusun
            today = date.today()
            age_groups = {'0-14': 0, '15-24': 0, '25-54': 0, '55-64': 0, '65+': 0}
            
            residents = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun
            )
            
            for resident in residents:
                age = today.year - resident.birth_date.year
                # Handle leap year edge case for February 29th
                try:
                    if resident.birth_date.replace(year=today.year) > today:
                        age -= 1
                except ValueError:
                    # Handle February 29th in non-leap year
                    if resident.birth_date.month == 2 and resident.birth_date.day == 29:
                        # Use February 28th for comparison
                        birth_date_this_year = resident.birth_date.replace(year=today.year, day=28)
                        if birth_date_this_year > today:
                            age -= 1
                    else:
                        # For other cases, just use the original logic
                        if resident.birth_date.replace(year=today.year) > today:
                            age -= 1
                    
                if age <= 14:
                    age_groups['0-14'] += 1
                elif age <= 24:
                    age_groups['15-24'] += 1
                elif age <= 54:
                    age_groups['25-54'] += 1
                elif age <= 64:
                    age_groups['55-64'] += 1
                else:
                    age_groups['65+'] += 1
            
            # Family count in this dusun
            family_count = Family.objects.filter(is_active=True, dusun=dusun).count()
            
            dusun_data.append({
                'id': dusun.id,
                'name': dusun.name,
                'code': dusun.code,
                'description': dusun.description or '',
                'area_size': float(dusun.area_size) if dusun.area_size else 0,
                'population': {
                    'total': total_count,
                    'male': male_count,
                    'female': female_count
                },
                'age_groups': age_groups,
                'family_count': family_count
            })
        
        return JsonResponse({
            'success': True,
            'data': dusun_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_search_residents(request):
    """API for searching residents with age filter 5-35 years - accessible by other apps"""
    try:
        query = request.GET.get('q', '').strip()
        dusun_id = request.GET.get('dusun_id')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Filter umur 5-35 tahun
        from datetime import date
        today = date.today()
        min_birth_date = today.replace(year=today.year - 35)  # 35 tahun ke bawah
        max_birth_date = today.replace(year=today.year - 5)   # 5 tahun ke atas
        
        # Base queryset dengan filter umur
        residents = Penduduk.objects.filter(
            is_active=True, 
            is_alive=True,
            birth_date__gte=min_birth_date,
            birth_date__lte=max_birth_date
        ).select_related('dusun', 'rt', 'rw', 'created_by', 'updated_by')
        
        # Apply search filter
        if query:
            residents = residents.filter(
                Q(name__icontains=query) |
                Q(nik__icontains=query) |
                Q(kk_number__icontains=query)
            )
        
        # Apply dusun filter
        if dusun_id:
            residents = residents.filter(dusun_id=dusun_id)
        
        # Order by name
        residents = residents.order_by('name')
        
        # Pagination
        paginator = Paginator(residents, per_page)
        page_obj = paginator.get_page(page)
        
        # Serialize data dengan informasi lengkap
        residents_data = []
        for resident in page_obj:
            residents_data.append({
                'id': resident.id,
                'nik': resident.nik,
                'name': resident.name,
                'gender': resident.get_gender_display(),
                'gender_code': resident.gender,
                'birth_date': resident.birth_date.strftime('%Y-%m-%d') if resident.birth_date else None,
                'birth_place': resident.birth_place or '',
                'age': resident.age if hasattr(resident, 'age') else (date.today().year - resident.birth_date.year if resident.birth_date else None),
                'dusun': resident.dusun.name if resident.dusun else '',
                'dusun_id': resident.dusun.id if resident.dusun else None,
                'rt_number': resident.rt_number or '',
                'rw_number': resident.rw_number or '',
                'address': resident.address or '',
                'full_address': resident.full_address or '',
                'phone': resident.phone_number or '',
                'occupation': resident.occupation or '',
                'education': resident.get_education_display() if resident.education else '',
                'education_code': resident.education or '',
                'education_level': resident.education or '',
                'religion': resident.religion or '',
                'religion_code': resident.religion or '',
                'marital_status': resident.get_marital_status_display() if resident.marital_status else '',
                'marital_status_code': resident.marital_status or '',
                'kk_number': resident.kk_number or '',
                'relationship_to_head': resident.get_relationship_to_head_display() if resident.relationship_to_head else '',
                'relationship_to_head_code': resident.relationship_to_head or '',
                'is_alive': resident.is_alive,
                'is_active': resident.is_active,
                'created_at': resident.created_at.isoformat() if resident.created_at else None,
                'updated_at': resident.updated_at.isoformat() if resident.updated_at else None
            })
        
        return JsonResponse({
            'success': True,
            'data': residents_data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_get_resident_detail(request, resident_id):
    """API to get detailed resident information - accessible by other apps"""
    try:
        resident = Penduduk.objects.get(id=resident_id, is_active=True)
        
        # Get family members if this person is family head
        family_members = []
        if resident.is_family_head:
            members = Penduduk.objects.filter(
                family_head=resident, is_active=True, is_alive=True
            ).exclude(id=resident.id)
            
            for member in members:
                family_members.append({
                    'id': member.id,
                    'name': member.name,
                    'relationship': member.relationship_to_head,
                    'age': member.age,
                    'gender': member.get_gender_display()
                })
        
        # Get disabilities
        disabilities = []
        for disability in resident.disabilities.filter(is_active=True):
            disabilities.append({
                'type': disability.disability_type.name,
                'severity': disability.get_severity_display(),
                'description': disability.description or ''
            })
        
        data = {
            'id': resident.id,
            'nik': resident.nik,
            'name': resident.name,
            'gender': resident.get_gender_display(),
            'birth_place': resident.birth_place,
            'birth_date': resident.birth_date.strftime('%Y-%m-%d'),
            'age': resident.age,
            'religion': resident.religion,
            'education': resident.get_education_display() if resident.education else '',
            'occupation': resident.occupation or '',
            'marital_status': resident.get_marital_status_display(),
            'blood_type': resident.blood_type or '',
            'phone_number': resident.phone_number or '',
            'mobile_number': resident.mobile_number or '',
            'email': resident.email or '',
            'address': resident.full_address,
            'dusun': resident.dusun.name,
            'lorong': resident.lorong.name if resident.lorong else '',
            'rt_number': resident.rt_number or '',
            'rw_number': resident.rw_number or '',
            'kk_number': resident.kk_number or '',
            'is_family_head': resident.is_family_head,
            'family_members': family_members,
            'disabilities': disabilities,
            'citizenship': resident.get_citizenship_display()
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Penduduk.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Penduduk tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_get_dusun_list(request):
    """API to get list of dusun - accessible by other apps"""
    try:
        dusun_list = []
        for dusun in Dusun.objects.filter(is_active=True).order_by('name'):
            population_count = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun
            ).count()
            
            dusun_list.append({
                'id': dusun.id,
                'name': dusun.name,
                'code': dusun.code,
                'population_count': population_count,
                'area_size': float(dusun.area_size) if dusun.area_size else 0
            })
        
        return JsonResponse({
            'success': True,
            'data': dusun_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def api_lorong_by_dusun(request):
    """API untuk mendapatkan daftar lorong berdasarkan dusun - accessible by other apps"""
    try:
        dusun_id = request.GET.get('dusun_id')
        if not dusun_id:
            return JsonResponse({'success': False, 'message': 'Parameter dusun_id diperlukan'})
        
        dusun = get_object_or_404(Dusun, id=dusun_id)
        lorong_list = Lorong.objects.filter(dusun=dusun, is_active=True).order_by('nama_lorong').values('id', 'nama_lorong')
        return JsonResponse({'success': True, 'data': list(lorong_list)})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@csrf_exempt
@require_http_methods(["GET"])
def api_public_population_overview(request):
    """Public API for population overview - no authentication required"""
    try:
        # Basic population counts
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        male_count = Penduduk.objects.filter(is_active=True, is_alive=True, gender='L').count()
        female_count = Penduduk.objects.filter(is_active=True, is_alive=True, gender='P').count()
        total_families = Family.objects.filter(is_active=True).count()
        
        # Age groups calculation
        today = date.today()
        age_groups = {
            '0-14': 0,
            '15-24': 0,
            '25-54': 0,
            '55-64': 0,
            '65+': 0
        }
        
        residents = Penduduk.objects.filter(is_active=True, is_alive=True)
        for resident in residents:
            age = today.year - resident.birth_date.year
            # Handle leap year edge case for February 29th
            try:
                if resident.birth_date.replace(year=today.year) > today:
                    age -= 1
            except ValueError:
                # Handle February 29th in non-leap year
                if resident.birth_date.month == 2 and resident.birth_date.day == 29:
                    # Use February 28th for comparison
                    birth_date_this_year = resident.birth_date.replace(year=today.year, day=28)
                    if birth_date_this_year > today:
                        age -= 1
                else:
                    # For other cases, just use the original logic
                    if resident.birth_date.replace(year=today.year) > today:
                        age -= 1
                
            if age <= 14:
                age_groups['0-14'] += 1
            elif age <= 24:
                age_groups['15-24'] += 1
            elif age <= 54:
                age_groups['25-54'] += 1
            elif age <= 64:
                age_groups['55-64'] += 1
            else:
                age_groups['65+'] += 1
        
        # Dusun statistics
        dusun_stats = []
        for dusun in Dusun.objects.filter(is_active=True):
            dusun_population = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun
            ).count()
            dusun_male = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='L'
            ).count()
            dusun_female = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='P'
            ).count()
            
            dusun_stats.append({
                'id': dusun.id,
                'name': dusun.name,
                'code': dusun.code,
                'total': dusun_population,
                'male': dusun_male,
                'female': dusun_female,
                'area_size': float(dusun.area_size) if dusun.area_size else 0
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_population': total_population,
                'male_population': male_count,
                'female_population': female_count,
                'total_families': total_families,
                'age_groups': age_groups,
                'dusun_stats': dusun_stats
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def api_public_population_demographics(request):
    """Public API for population demographics - no authentication required"""
    try:
        # Age distribution with percentages
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        
        if total_population == 0:
            return JsonResponse({
                'success': True,
                'data': {
                    'age_distribution': {},
                    'marital_status': {},
                    'religion': {},
                    'disability': {},
                    'blood_type': {}
                }
            })
        
        # Age distribution
        today = date.today()
        age_distribution = {
            '0-14': 0,
            '15-24': 0,
            '25-54': 0,
            '55-64': 0,
            '65+': 0
        }
        
        residents = Penduduk.objects.filter(is_active=True, is_alive=True)
        for resident in residents:
            age = today.year - resident.birth_date.year
            # Handle leap year edge case for February 29th
            try:
                if resident.birth_date.replace(year=today.year) > today:
                    age -= 1
            except ValueError:
                # Handle February 29th in non-leap year
                if resident.birth_date.month == 2 and resident.birth_date.day == 29:
                    # Use February 28th for comparison
                    birth_date_this_year = resident.birth_date.replace(year=today.year, day=28)
                    if birth_date_this_year > today:
                        age -= 1
                else:
                    # For other cases, just use the original logic
                    if resident.birth_date.replace(year=today.year) > today:
                        age -= 1
                
            if age <= 14:
                age_distribution['0-14'] += 1
            elif age <= 24:
                age_distribution['15-24'] += 1
            elif age <= 54:
                age_distribution['25-54'] += 1
            elif age <= 64:
                age_distribution['55-64'] += 1
            else:
                age_distribution['65+'] += 1
        
        # Convert to percentages
        for key in age_distribution:
            count = age_distribution[key]
            percentage = (count / total_population * 100) if total_population > 0 else 0
            age_distribution[key] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        # Marital status distribution
        marital_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True
        ).values('marital_status').annotate(count=Count('id'))
        
        marital_status = {}
        for stat in marital_stats:
            status = stat['marital_status']
            count = stat['count']
            percentage = (count / total_population * 100) if total_population > 0 else 0
            marital_status[status] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        # Religion distribution
        religion_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True
        ).values('religion').annotate(count=Count('id'))
        
        religion = {}
        for stat in religion_stats:
            rel = stat['religion']
            count = stat['count']
            percentage = (count / total_population * 100) if total_population > 0 else 0
            religion[rel] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        # Disability statistics
        disability_count = DisabilitasData.objects.filter(
            is_active=True, penduduk__is_active=True, penduduk__is_alive=True
        ).values('penduduk').distinct().count()
        
        disability = {
            'with_disability': {
                'count': disability_count,
                'percentage': round((disability_count / total_population * 100), 1) if total_population > 0 else 0
            },
            'without_disability': {
                'count': total_population - disability_count,
                'percentage': round(((total_population - disability_count) / total_population * 100), 1) if total_population > 0 else 0
            }
        }
        
        # Blood type distribution
        blood_type_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True, blood_type__isnull=False
        ).exclude(blood_type='').values('blood_type').annotate(count=Count('id'))
        
        blood_type = {}
        total_with_blood_type = sum([stat['count'] for stat in blood_type_stats])
        
        for stat in blood_type_stats:
            bt = stat['blood_type']
            count = stat['count']
            percentage = (count / total_with_blood_type * 100) if total_with_blood_type > 0 else 0
            blood_type[bt] = {
                'count': count,
                'percentage': round(percentage, 1)
            }
        
        return JsonResponse({
            'success': True,
            'data': {
                'age_distribution': age_distribution,
                'marital_status': marital_status,
                'religion': religion,
                'disability': disability,
                'blood_type': blood_type
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def api_public_statistics(request):
    """Comprehensive public statistics API for homepage and public pages"""
    try:
        # Population statistics
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        male_population = Penduduk.objects.filter(is_active=True, is_alive=True, gender='L').count()
        female_population = Penduduk.objects.filter(is_active=True, is_alive=True, gender='P').count()
        total_families = Family.objects.filter(is_active=True).count()
        
        # News statistics
        news_stats = {
            'total': 0,
            'this_month': 0,
            'this_year': 0
        }
        
        if NEWS_AVAILABLE:
            news_stats['total'] = News.objects.filter(status='published').count()
            
            # This month
            this_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            news_stats['this_month'] = News.objects.filter(
                status='published', 
                created_at__gte=this_month
            ).count()
            
            # This year
            this_year = timezone.now().replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            news_stats['this_year'] = News.objects.filter(
                status='published', 
                created_at__gte=this_year
            ).count()
        
        # Business/UMKM statistics
        business_stats = {
            'total': 0,
            'active': 0,
            'categories': {}
        }
        
        if BUSINESS_AVAILABLE:
            business_stats['total'] = UMKM.objects.filter(is_active=True).count()
            business_stats['active'] = UMKM.objects.filter(is_active=True, status='ACTIVE').count()
            
            # Business categories
            categories = UMKM.objects.filter(is_active=True).values('business_type').annotate(
                count=Count('id')
            ).order_by('-count')[:5]
            
            for cat in categories:
                business_stats['categories'][cat['business_type']] = cat['count']
        
        # Tourism statistics
        tourism_stats = {
            'total': 0,
            'featured': 0
        }
        
        if TOURISM_AVAILABLE:
            tourism_stats['total'] = TourismLocation.objects.filter(is_active=True).count()
            tourism_stats['featured'] = TourismLocation.objects.filter(
                is_active=True, 
                featured=True
            ).count()
        
        # Age distribution for charts
        today = date.today()
        age_groups = {
            '0-14': 0,
            '15-24': 0,
            '25-54': 0,
            '55-64': 0,
            '65+': 0
        }
        
        residents = Penduduk.objects.filter(is_active=True, is_alive=True)
        for resident in residents:
            age = today.year - resident.birth_date.year
            # Handle leap year edge case for February 29th
            try:
                if resident.birth_date.replace(year=today.year) > today:
                    age -= 1
            except ValueError:
                # Handle February 29th in non-leap year
                if resident.birth_date.month == 2 and resident.birth_date.day == 29:
                    # Use February 28th for comparison
                    birth_date_this_year = resident.birth_date.replace(year=today.year, day=28)
                    if birth_date_this_year > today:
                        age -= 1
                else:
                    # For other cases, just use the original logic
                    if resident.birth_date.replace(year=today.year) > today:
                        age -= 1
                
            if age <= 14:
                age_groups['0-14'] += 1
            elif age <= 24:
                age_groups['15-24'] += 1
            elif age <= 54:
                age_groups['25-54'] += 1
            elif age <= 64:
                age_groups['55-64'] += 1
            else:
                age_groups['65+'] += 1
        
        # Education summary
        education_summary = {
            'higher_education': 0,  # D4/S1/S2/S3
            'secondary_education': 0,  # SLTA
            'primary_education': 0,  # SD/SLTP
            'no_education': 0
        }
        
        for resident in residents:
            if resident.education in ['D4_S1', 'S2', 'S3']:
                education_summary['higher_education'] += 1
            elif resident.education == 'SLTA':
                education_summary['secondary_education'] += 1
            elif resident.education in ['TAMAT_SD', 'SLTP']:
                education_summary['primary_education'] += 1
            elif resident.education in ['TIDAK_BELUM_SEKOLAH', 'BELUM_TAMAT_SD']:
                education_summary['no_education'] += 1
        
        # Occupation summary
        occupation_summary = {}
        occupation_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True, occupation__isnull=False
        ).exclude(occupation='').values('occupation').annotate(count=Count('id')).order_by('-count')[:10]
        
        for stat in occupation_stats:
            occupation_summary[stat['occupation']] = stat['count']
        
        return JsonResponse({
            'success': True,
            'data': {
                'population': {
                    'total': total_population,
                    'male': male_population,
                    'female': female_population,
                    'families': total_families
                },
                'news': news_stats,
                'business': business_stats,
                'tourism': tourism_stats,
                'demographics': {
                    'age_groups': age_groups,
                    'education_summary': education_summary,
                    'occupation_summary': occupation_summary
                }
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_dashboard_statistics(request):
    """Dashboard statistics for admin and public dashboard"""
    try:
        # Population growth (last 12 months)
        population_growth = []
        for i in range(12):
            month_date = timezone.now() - timedelta(days=30*i)
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            if i == 0:
                month_end = timezone.now()
            else:
                next_month = month_date + timedelta(days=32)
                month_end = next_month.replace(day=1) - timedelta(days=1)
            
            count = Penduduk.objects.filter(
                is_active=True,
                created_at__gte=month_start,
                created_at__lte=month_end
            ).count()
            
            population_growth.append({
                'month': month_start.strftime('%Y-%m'),
                'month_name': month_start.strftime('%b %Y'),
                'new_residents': count
            })
        
        population_growth.reverse()  # Oldest first
        
        # Recent activities
        recent_activities = []
        
        # Recent residents (last 30 days)
        recent_residents = Penduduk.objects.filter(
            is_active=True,
            created_at__gte=timezone.now() - timedelta(days=30)
        ).order_by('-created_at')[:5]
        
        for resident in recent_residents:
            recent_activities.append({
                'type': 'new_resident',
                'title': f'Penduduk baru: {resident.name}',
                'date': resident.created_at.strftime('%Y-%m-%d'),
                'description': f'Ditambahkan ke Dusun {resident.dusun.name}'
            })
        
        # Recent news (if available)
        if NEWS_AVAILABLE:
            recent_news = News.objects.filter(
                status='published',
                created_at__gte=timezone.now() - timedelta(days=30)
            ).order_by('-created_at')[:3]
            
            for news in recent_news:
                recent_activities.append({
                    'type': 'news',
                    'title': f'Berita: {news.title}',
                    'date': news.created_at.strftime('%Y-%m-%d'),
                    'description': news.excerpt or 'Berita terbaru'
                })
        
        # Family statistics
        family_stats = {
            'total_families': Family.objects.filter(is_active=True).count(),
            'average_family_size': 0,
            'largest_family_size': 0,
            'family_status_distribution': {}
        }
        
        families = Family.objects.filter(is_active=True)
        if families.exists():
            family_sizes = [f.total_members for f in families]
            family_stats['average_family_size'] = round(sum(family_sizes) / len(family_sizes), 1)
            family_stats['largest_family_size'] = max(family_sizes)
            
            # Family status distribution
            status_dist = families.values('family_status').annotate(count=Count('id'))
            for stat in status_dist:
                family_stats['family_status_distribution'][stat['family_status']] = stat['count']
        
        return JsonResponse({
            'success': True,
            'data': {
                'population_growth': population_growth,
                'recent_activities': recent_activities,
                'family_stats': family_stats
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_public_news(request):
    """Public news API for homepage"""
    try:
        if not NEWS_AVAILABLE:
            return JsonResponse({
                'success': True,
                'data': []
            })
        
        # Get latest 3 news items
        latest_news = News.objects.filter(
            status='published'
        ).order_by('-created_at')[:3]
        
        news_data = []
        for news in latest_news:
            news_data.append({
                'id': news.id,
                'title': news.title,
                'excerpt': news.excerpt or news.content[:150] + '...' if len(news.content) > 150 else news.content,
                'created_at': news.created_at.strftime('%d %B %Y'),
                'author': news.author.get_full_name() if news.author else 'Admin',
                'category': news.category.name if news.category else 'Umum',
                'url': f'/berita/{news.slug}/' if hasattr(news, 'slug') else f'/berita/{news.id}/'
            })
        
        return JsonResponse({
            'success': True,
            'data': news_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_public_business(request):
    """Public business/UMKM API for homepage"""
    try:
        if not BUSINESS_AVAILABLE:
            return JsonResponse({
                'success': True,
                'data': []
            })
        
        # Get featured or active businesses
        businesses = UMKM.objects.filter(
            is_active=True,
            status='ACTIVE'
        ).order_by('-created_at')[:6]
        
        business_data = []
        for business in businesses:
            business_data.append({
                'id': business.id,
                'name': business.name,
                'description': business.description[:100] + '...' if len(business.description) > 100 else business.description,
                'business_type': business.business_type,
                'owner_name': business.owner_name,
                'phone': business.phone_number,
                'address': business.address,
                'created_at': business.created_at.strftime('%d %B %Y')
            })
        
        return JsonResponse({
            'success': True,
            'data': business_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_public_tourism(request):
    """Public tourism API for homepage"""
    try:
        if not TOURISM_AVAILABLE:
            return JsonResponse({
                'success': True,
                'data': []
            })
        
        # Get featured tourism locations
        tourism_locations = TourismLocation.objects.filter(
            is_active=True,
            featured=True
        ).order_by('-created_at')[:4]
        
        tourism_data = []
        for location in tourism_locations:
            tourism_data.append({
                'id': location.id,
                'name': location.title,
                'description': location.short_description[:100] + '...' if len(location.short_description) > 100 else location.short_description,
                'location': location.address,
                'category': location.category.name if location.category else '',
                'rating': getattr(location, 'average_rating', 0),
                'image_url': location.main_image.url if location.main_image else None
            })
        
        return JsonResponse({
            'success': True,
            'data': tourism_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# REFERENCE STATISTICS API - FOR EXTERNAL WEBSITE INTEGRATION
# =============================================================================

@require_http_methods(["GET"])
@login_required
@cache_page(60 * 5)  # Cache for 5 minutes
def api_reference_statistics(request):
    """
    Comprehensive reference statistics API for external website integration
    
    Authentication: Required (Login required)
    Cache: 5 minutes
    Response: JSON with complete statistics data
    
    Usage:
    GET /references/api/reference/statistics/
    Headers: Authorization: Bearer <token> (if using token auth)
    
    Response Format:
    {
        "success": true,
        "data": {
            "population": {...},
            "dusun": {...},
            "education": {...},
            "disability": {...},
            "family": {...},
            "demographics": {...}
        },
        "timestamp": "2024-01-01T00:00:00Z",
        "version": "1.0"
    }
    """
    try:
        # Get chart data for the last 12 months
        chart_data = get_chart_data()
        
        # Get current statistics (only active records)
        total_penduduk = Penduduk.objects.filter(is_active=True).count()
        total_keluarga = Keluarga.objects.filter(is_active=True).count()
        total_dusun = Dusun.objects.filter(is_active=True).count()
        total_pelajar = Pelajar.objects.filter(is_active=True).count()
        
        # Gender distribution
        total_laki_laki = Penduduk.objects.filter(gender='L', is_active=True).count()
        total_perempuan = Penduduk.objects.filter(gender='P', is_active=True).count()
        
        # Age group distribution
        today = date.today()
        age_groups = {
            'balita': Penduduk.objects.filter(
                birth_date__gte=today - timedelta(days=5*365), is_active=True
            ).count(),
            'anak': Penduduk.objects.filter(
                birth_date__gte=today - timedelta(days=17*365),
                birth_date__lt=today - timedelta(days=5*365), is_active=True
            ).count(),
            'dewasa': Penduduk.objects.filter(
                birth_date__gte=today - timedelta(days=60*365),
                birth_date__lt=today - timedelta(days=17*365)
            ).count(),
            'lansia': Penduduk.objects.filter(
                birth_date__lt=today - timedelta(days=60*365), is_active=True
            ).count(),
        }
        
        # Education statistics
        education_queryset = Penduduk.objects.filter(
            education__isnull=False, is_active=True
        ).exclude(education='').values('education').annotate(
            count=Count('id')
        ).order_by('education')
        education_stats = list(education_queryset)
        
        # Dusun population distribution
        dusun_stats = Dusun.objects.filter(is_active=True).annotate(
            current_population=Count('residents', filter=Q(residents__is_active=True))
        ).order_by('-current_population')
        
        return JsonResponse({
            'success': True,
            'chart_data': {
                'labels': chart_data['labels'],
                'penduduk_data': chart_data['penduduk_data'],
                'keluarga_data': chart_data['keluarga_data'],
                'dusun_data': chart_data['dusun_data'],
                'total_penduduk': total_penduduk,
                'total_keluarga': total_keluarga,
                'total_dusun': total_dusun,
                'total_pelajar': total_pelajar,
            },
            'demographics': {
                'gender': {
                    'male': total_laki_laki,
                    'female': total_perempuan,
                },
                'age_groups': age_groups,
                'education': education_stats,
                'dusun_distribution': list(dusun_stats.values('name', 'current_population')),
            },
            'timestamp': timezone.now().isoformat(),
            'version': '1.0'
        }, content_type='application/json')
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# ADMIN PANEL INTEGRATION
# =============================================================================


@require_http_methods(["GET"])
@login_required
@cache_page(60 * 10)  # Cache for 10 minutes
def api_reference_summary(request):
    """
    Quick summary statistics API for external website integration
    
    Authentication: Required (Login required)
    Cache: 10 minutes
    Response: JSON with summary statistics only
    
    Usage:
    GET /references/api/reference/summary/
    Headers: Authorization: Bearer <token> (if using token auth)
    
    Response Format:
    {
        "success": true,
        "data": {
            "total_population": 1500,
            "total_dusun": 5,
            "total_families": 400,
            "total_with_disability": 25,
            "education_levels": 8,
            "religion_types": 6
        },
        "timestamp": "2024-01-01T00:00:00Z"
    }
    """
    try:
        # Quick counts only
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        total_dusun = Dusun.objects.filter(is_active=True).count()
        total_families = Family.objects.filter(is_active=True).count()
        total_with_disability = DisabilitasData.objects.filter(
            is_active=True, 
            penduduk__is_active=True, 
            penduduk__is_alive=True
        ).values('penduduk').distinct().count()
        
        education_levels = Penduduk.objects.filter(
            is_active=True, is_alive=True, education__isnull=False
        ).exclude(education='').values('education').distinct().count()
        
        religion_types = Penduduk.objects.filter(
            is_active=True, is_alive=True
        ).values('religion').distinct().count()
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_population': total_population,
                'total_dusun': total_dusun,
                'total_families': total_families,
                'total_with_disability': total_with_disability,
                'education_levels': education_levels,
                'religion_types': religion_types
            },
            'timestamp': timezone.now().isoformat(),
            'version': '1.0'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_reference_dusun_population(request):
    """
    Dusun population distribution API
    
    Authentication: Required (Login required)
    Response: JSON with dusun population data
    
    Usage:
    GET /references/api/reference/dusun-population/
    Headers: Authorization: Bearer <token> (if using token auth)
    
    Response Format:
    {
        "success": true,
        "data": [
            {
                "dusun_name": "Dusun A",
                "population": 300,
                "male": 150,
                "female": 150,
                "families": 75
            }
        ]
    }
    """
    try:
        dusun_data = []
        
        for dusun in Dusun.objects.filter(is_active=True).order_by('name'):
            total_pop = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun
            ).count()
            
            male_pop = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='L'
            ).count()
            
            female_pop = Penduduk.objects.filter(
                is_active=True, is_alive=True, dusun=dusun, gender='P'
            ).count()
            
            family_count = Family.objects.filter(is_active=True, dusun=dusun).count()
            
            dusun_data.append({
                'dusun_name': dusun.name,
                'population': total_pop,
                'male': male_pop,
                'female': female_pop,
                'families': family_count
            })
        
        return JsonResponse({
            'success': True,
            'data': dusun_data,
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_reference_education_distribution(request):
    """
    Education level distribution API
    
    Authentication: Required (Login required)
    Response: JSON with education distribution data
    
    Usage:
    GET /references/api/reference/education-distribution/
    Headers: Authorization: Bearer <token> (if using token auth)
    """
    try:
        education_stats = Penduduk.objects.filter(
            is_active=True, is_alive=True, education__isnull=False
        ).exclude(education='').values('education').annotate(count=Count('id')).order_by('-count')
        
        total_with_education = sum([stat['count'] for stat in education_stats])
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        
        education_data = []
        for stat in education_stats:
            percentage = round((stat['count'] / total_with_education * 100), 2) if total_with_education > 0 else 0
            education_data.append({
                'education_level': stat['education'],
                'count': stat['count'],
                'percentage': percentage
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_with_education': total_with_education,
                'total_without_education': total_population - total_with_education,
                'distribution': education_data
            },
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_reference_disability_breakdown(request):
    """
    Disability breakdown by type and severity API
    
    Authentication: Required (Login required)
    Response: JSON with disability breakdown data
    
    Usage:
    GET /references/api/reference/disability-breakdown/
    Headers: Authorization: Bearer <token> (if using token auth)
    """
    try:
        # By disability type
        disability_by_type = DisabilitasData.objects.filter(
            is_active=True,
            penduduk__is_active=True,
            penduduk__is_alive=True
        ).values('disability_type__name').annotate(count=Count('id')).order_by('-count')
        
        # By severity
        disability_by_severity = DisabilitasData.objects.filter(
            is_active=True,
            penduduk__is_active=True,
            penduduk__is_alive=True
        ).values('severity').annotate(count=Count('id')).order_by('-count')
        
        # Total counts
        total_with_disability = DisabilitasData.objects.filter(
            is_active=True,
            penduduk__is_active=True,
            penduduk__is_alive=True
        ).values('penduduk').distinct().count()
        
        total_population = Penduduk.objects.filter(is_active=True, is_alive=True).count()
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_with_disability': total_with_disability,
                'total_without_disability': total_population - total_with_disability,
                'disability_rate': round((total_with_disability / total_population * 100), 2) if total_population > 0 else 0,
                'by_type': [{'type': item['disability_type__name'], 'count': item['count']} for item in disability_by_type],
                'by_severity': [{'severity': item['severity'], 'count': item['count']} for item in disability_by_severity]
            },
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=500)


@csrf_exempt
def api_documentation(request):
    """
    API Documentation page for reference statistics API
    
    Usage:
    GET /references/api/documentation/
    """
    context = {
        'page_title': 'Reference Statistics API Documentation',
        'api_endpoints': [
            {
                'endpoint': '/references/api/reference/statistics/',
                'method': 'GET',
                'description': 'Comprehensive reference statistics',
                'auth_required': True,
                'cache': '5 minutes',
                'response_fields': [
                    'population (total, male, female, gender_ratio)',
                    'dusun (total, with_population, details)',
                    'education (total_with_education, no_education, breakdown)',
                    'disability (total_with_disability, disability_rate, by_type)',
                    'family (total_families, average_family_size, by_status)',
                    'demographics (age_groups, religion, marital_status, blood_type, occupation)'
                ]
            },
            {
                'endpoint': '/references/api/reference/summary/',
                'method': 'GET',
                'description': 'Quick summary statistics',
                'auth_required': True,
                'cache': '10 minutes',
                'response_fields': [
                    'total_population',
                    'total_dusun',
                    'total_families',
                    'total_with_disability',
                    'education_levels',
                    'religion_types'
                ]
            },
            {
                'endpoint': '/references/api/reference/dusun-population/',
                'method': 'GET',
                'description': 'Dusun population distribution',
                'auth_required': True,
                'cache': 'None',
                'response_fields': [
                    'dusun_name',
                    'population',
                    'male',
                    'female',
                    'families'
                ]
            },
            {
                'endpoint': '/references/api/reference/education-distribution/',
                'method': 'GET',
                'description': 'Education level distribution',
                'auth_required': True,
                'cache': 'None',
                'response_fields': [
                    'total_with_education',
                    'total_without_education',
                    'distribution (education_level, count, percentage)'
                ]
            },
            {
                'endpoint': '/references/api/reference/disability-breakdown/',
                'method': 'GET',
                'description': 'Disability breakdown by type and severity',
                'auth_required': True,
                'cache': 'None',
                'response_fields': [
                    'total_with_disability',
                    'total_without_disability',
                    'disability_rate',
                    'by_type',
                    'by_severity'
                ]
            }
        ]
    }
    return render(request, 'references/api_documentation.html', context)


# ============= ADMIN PANEL INTEGRATION =============

# Admin Panel Dashboard for References
@login_required
def admin_panel_dashboard(request):
    """Dashboard admin panel untuk aplikasi references"""
    # Check if user has references permission
    if not request.user.has_menu_permission('references', 'view'):
        from django.contrib import messages
        messages.error(request, 'Anda tidak memiliki izin untuk mengakses halaman ini.')
        return redirect('admin_panel:dashboard')
    from django.db.models import Count, Avg, Sum
    from datetime import datetime, timedelta
    
    # Basic statistics (only active records)
    total_penduduk = Penduduk.objects.filter(is_active=True).count()
    total_dusun = Dusun.objects.filter(is_active=True).count()
    total_lorong = Lorong.objects.filter(is_active=True).count()
    total_families = Family.objects.filter(is_active=True).count()
    total_pelajar = Pelajar.objects.filter(is_active=True).count()
    total_disabilitas = DisabilitasData.objects.filter(is_active=True).count()
    total_disabilitas_types = DisabilitasType.objects.filter(is_active=True).count()
    total_religion_refs = ReligionReference.objects.filter(is_active=True).count()
    total_keluarga = Keluarga.objects.filter(is_active=True).count()
    
    # Gender distribution
    gender_stats = Penduduk.objects.filter(is_active=True).values('gender').annotate(
        count=Count('id')
    ).order_by('gender')
    
    # Gender counts for template
    total_laki_laki = Penduduk.objects.filter(gender='L', is_active=True).count()
    total_perempuan = Penduduk.objects.filter(gender='P', is_active=True).count()
    
    # Age group distribution
    today = date.today()
    age_groups = {
        'balita': Penduduk.objects.filter(
            birth_date__gte=today - timedelta(days=5*365), is_active=True
        ).count(),
        'anak': Penduduk.objects.filter(
            birth_date__gte=today - timedelta(days=17*365),
            birth_date__lt=today - timedelta(days=5*365), is_active=True
        ).count(),
        'dewasa': Penduduk.objects.filter(
            birth_date__gte=today - timedelta(days=60*365),
            birth_date__lt=today - timedelta(days=17*365)
        ).count(),
        'lansia': Penduduk.objects.filter(
            birth_date__lt=today - timedelta(days=60*365), is_active=True
        ).count(),
    }
    
    # Age counts for template
    total_anak = age_groups['anak'] + age_groups['balita']
    total_dewasa = age_groups['dewasa']
    total_lansia = age_groups['lansia']
    
    # Education statistics
    education_queryset = Penduduk.objects.filter(
        education__isnull=False
    ).exclude(education='').values('education').annotate(
        count=Count('id')
    ).order_by('education')
    education_stats = json.dumps(list(education_queryset))
    
    # Occupation statistics
    occupation_queryset = Penduduk.objects.filter(
        occupation__isnull=False
    ).exclude(occupation='').values('occupation').annotate(
        count=Count('id')
    ).order_by('occupation')
    occupation_stats = json.dumps(list(occupation_queryset))
    
    # Religion statistics
    religion_queryset = Penduduk.objects.values('religion').annotate(
        count=Count('id')
    ).order_by('religion')
    religion_stats = json.dumps(list(religion_queryset))
    
    # Marital status statistics
    marital_queryset = Penduduk.objects.values('marital_status').annotate(
        count=Count('id')
    ).order_by('marital_status')
    marital_stats = json.dumps(list(marital_queryset))
    
    # Education level statistics for Pelajar
    pelajar_jenjang_queryset = Pelajar.objects.filter(is_active=True).values('jenjang').annotate(
        count=Count('id')
    ).order_by('jenjang')
    pelajar_jenjang_stats = json.dumps(list(pelajar_jenjang_queryset))
    
    # Disabilitas type statistics
    disabilitas_type_queryset = DisabilitasData.objects.filter(is_active=True).values('disability_type__name').annotate(
        count=Count('id')
    ).order_by('disability_type__name')
    disabilitas_type_stats = json.dumps(list(disabilitas_type_queryset))
    
    # Dusun population distribution
    dusun_stats = Dusun.objects.annotate(
        current_population=Count('residents')
    ).order_by('-current_population')
    
    # Family structure statistics (by member count)
    family_structure_queryset = Family.objects.filter(
        total_members__gt=0, is_active=True
    ).values('total_members').annotate(
        count=Count('id')
    ).order_by('total_members')
    family_structure_stats = json.dumps(list(family_structure_queryset))
    
    # Recent activities (last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_penduduk = Penduduk.objects.filter(
        created_at__gte=thirty_days_ago, is_active=True
    ).count() if hasattr(Penduduk, 'created_at') else 0
    
    context = {
        'total_penduduk': total_penduduk,
        'total_dusun': total_dusun,
        'total_lorong': total_lorong,
        'total_keluarga': total_families,
        'total_pelajar': total_pelajar,
        'total_disabilitas': total_disabilitas,
        'total_disabilitas_types': total_disabilitas_types,
        'total_religion_refs': total_religion_refs,
        'gender_stats': gender_stats,
        'age_groups': age_groups,
        'education_stats': education_stats,
        'occupation_stats': occupation_stats,
        'religion_stats': religion_stats,
        'marital_stats': marital_stats,
        'pelajar_jenjang_stats': pelajar_jenjang_stats,
        'disabilitas_type_stats': disabilitas_type_stats,
        'dusun_stats': dusun_stats,
        'family_structure_stats': family_structure_stats,
        'recent_penduduk': recent_penduduk,
        # Template variables for dashboard
        'total_laki_laki': total_laki_laki,
        'total_perempuan': total_perempuan,
        'laki_laki': total_laki_laki,
        'perempuan': total_perempuan,
        'total_anak': total_anak,
        'total_dewasa': total_dewasa,
        'total_lansia': total_lansia,
        'anak': total_anak,
        'dewasa': total_dewasa,
        'lansia': total_lansia,
    }
    return render(request, 'admin_panel/references/dashboard.html', context)

@login_required
def api_references_dashboard_data(request):
    """API endpoint untuk data real-time dashboard references"""
    try:
        # Basic statistics (only active records)
        total_penduduk = Penduduk.objects.filter(is_active=True).count()
        total_dusun = Dusun.objects.filter(is_active=True).count()
        total_lorong = Lorong.objects.filter(is_active=True).count()
        total_families = Family.objects.filter(is_active=True).count()
        total_pelajar = Pelajar.objects.filter(is_active=True).count()
        total_disabilitas = DisabilitasData.objects.filter(is_active=True).count()
        total_keluarga = Keluarga.objects.filter(is_active=True).count()
        total_rw = RW.objects.filter(is_active=True).count()
        total_rt = RT.objects.filter(is_active=True).count()
        
        # Gender distribution
        total_laki_laki = Penduduk.objects.filter(gender='L', is_active=True).count()
        total_perempuan = Penduduk.objects.filter(gender='P', is_active=True).count()
        
        # Age group distribution
        today = date.today()
        age_groups = {
            'balita': Penduduk.objects.filter(
                birth_date__gte=today - timedelta(days=5*365), is_active=True
            ).count(),
            'anak': Penduduk.objects.filter(
                birth_date__gte=today - timedelta(days=17*365),
                birth_date__lt=today - timedelta(days=5*365), is_active=True
            ).count(),
            'dewasa': Penduduk.objects.filter(
                birth_date__gte=today - timedelta(days=60*365),
                birth_date__lt=today - timedelta(days=17*365)
            ).count(),
            'lansia': Penduduk.objects.filter(
                birth_date__lt=today - timedelta(days=60*365), is_active=True
            ).count(),
        }
        
        # Education statistics
        education_queryset = Penduduk.objects.filter(
            education__isnull=False, is_active=True
        ).exclude(education='').values('education').annotate(
            count=Count('id')
        ).order_by('education')
        education_stats = list(education_queryset)
        
        # Occupation statistics
        occupation_queryset = Penduduk.objects.filter(
            occupation__isnull=False, is_active=True
        ).exclude(occupation='').values('occupation').annotate(
            count=Count('id')
        ).order_by('occupation')
        occupation_stats = list(occupation_queryset)
        
        # Family structure statistics
        family_structure_queryset = Family.objects.filter(
            total_members__gt=0
        ).values('total_members').annotate(
            count=Count('id')
        ).order_by('total_members')
        family_structure_stats = list(family_structure_queryset)
        
        # Dusun population distribution
        dusun_stats = Dusun.objects.filter(is_active=True).annotate(
            current_population=Count('residents', filter=Q(residents__is_active=True))
        ).order_by('-current_population')
        
        # Prepare chart data
        dusun_labels = [dusun.name for dusun in dusun_stats]
        dusun_data = [dusun.current_population for dusun in dusun_stats]
        
        # Get chart data for last 6 months
        chart_data = get_chart_data()
        
        # Get last update time
        last_penduduk = Penduduk.objects.filter(is_active=True).order_by('-updated_at').first()
        last_family = Family.objects.filter(is_active=True).order_by('-updated_at').first()
        last_update = timezone.now()
        
        if last_penduduk and last_family:
            last_update = max(last_penduduk.updated_at, last_family.updated_at)
        elif last_penduduk:
            last_update = last_penduduk.updated_at
        elif last_family:
            last_update = last_family.updated_at
        
        # Calculate data integrity
        data_integrity = 100
        if total_penduduk > 0:
            missing_data = Penduduk.objects.filter(
                Q(nik__isnull=True) | Q(nik='') |
                Q(name__isnull=True) | Q(name='') |
                Q(birth_date__isnull=True),
                is_active=True
            ).count()
            data_integrity = max(0, 100 - (missing_data / total_penduduk * 100))
        
        return JsonResponse({
            'success': True,
            'total_penduduk': total_penduduk,
            'total_keluarga': total_keluarga,
            'total_families': total_families,
            'total_dusun': total_dusun,
            'total_pelajar': total_pelajar,
            'total_lorong': total_lorong,
            'total_disabilitas': total_disabilitas,
            'total_rw': total_rw,
            'total_rt': total_rt,
            'total_laki_laki': total_laki_laki,
            'total_perempuan': total_perempuan,
            'age_groups': age_groups,
            'education_stats': education_stats,
            'occupation_stats': occupation_stats,
            'family_structure_stats': family_structure_stats,
            'chart_data': chart_data,
            'last_update': last_update.isoformat(),
            'data_integrity': round(data_integrity, 1),
            'timestamp': timezone.now().isoformat()
        }, content_type='application/json')
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# Penduduk Management Views
@login_required
def admin_panel_penduduk_list(request):
    """List semua data penduduk"""
    # Check if user has references permission
    if not request.user.has_menu_permission('references', 'view'):
        from django.contrib import messages
        messages.error(request, 'Anda tidak memiliki izin untuk mengakses halaman ini.')
        return redirect('admin_panel:dashboard')

    search_query = request.GET.get('search', '')
    
    # Get penduduk data
    penduduk_list = Penduduk.objects.all()
    
    if search_query:
        penduduk_list = penduduk_list.filter(
            Q(nama__icontains=search_query) |
            Q(nik__icontains=search_query) |
            Q(alamat__icontains=search_query)
        )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(penduduk_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_penduduk': penduduk_list.count(),
    }
    
    return render(request, 'admin_panel/references/penduduk_list.html', context)

def get_chart_data():
    """Generate chart data for the last 12 months"""
    try:
        from datetime import datetime, timedelta
        import calendar
        
        # Get data for last 12 months
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=365)
        
        # Generate month labels
        months = []
        penduduk_data = []
        keluarga_data = []
        dusun_data = []
        
        current_date = start_date
        while current_date <= end_date:
            # Get first day of month
            month_start = current_date.replace(day=1)
            # Get last day of month
            if current_date.month == 12:
                month_end = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)
            
            # Count penduduk created in this month
            penduduk_count = Penduduk.objects.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end,
                is_active=True
            ).count()
            
            # Count keluarga created in this month
            keluarga_count = Keluarga.objects.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end,
                is_active=True
            ).count()
            
            # Count dusun created in this month
            dusun_count = Dusun.objects.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end,
                is_active=True
            ).count()
            
            months.append(calendar.month_name[current_date.month][:3])
            penduduk_data.append(penduduk_count)
            keluarga_data.append(keluarga_count)
            dusun_data.append(dusun_count)
            
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        return {
            'labels': months[-12:],  # Last 12 months
            'penduduk_data': penduduk_data[-12:],
            'keluarga_data': keluarga_data[-12:],
            'dusun_data': dusun_data[-12:],
        }
        
    except Exception as e:
        # Return default data if error
        return {
            'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des'],
            'penduduk_data': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            'keluarga_data': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            'dusun_data': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        }

# Penduduk Management Views
@login_required
def penduduk_list(request):
    """View untuk menampilkan daftar penduduk"""
    try:
        # Get search and filter parameters
        search_query = request.GET.get('search', '')
        gender_filter = request.GET.get('gender', '')
        dusun_filter = request.GET.get('dusun', '')
        marital_status_filter = request.GET.get('marital_status', '')
        
        # Get penduduk data with filters
        penduduk_queryset = Penduduk.objects.filter(is_active=True).select_related('dusun')
        
        # Apply filters
        if search_query:
            penduduk_queryset = penduduk_queryset.filter(
                Q(name__icontains=search_query) |
                Q(nik__icontains=search_query)
            )
        
        if gender_filter:
            penduduk_queryset = penduduk_queryset.filter(gender=gender_filter)
        
        if dusun_filter:
            penduduk_queryset = penduduk_queryset.filter(dusun_id=dusun_filter)
        
        if marital_status_filter:
            penduduk_queryset = penduduk_queryset.filter(marital_status=marital_status_filter)
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(penduduk_queryset, 20)  # 20 items per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Get dusun options for filter
        dusun_options = Dusun.objects.filter(is_active=True).order_by('name')
        
        # Debug info
        total_count = Penduduk.objects.count()
        print(f"Total penduduk in DB: {total_count}")
        print(f"Filtered penduduk count: {penduduk_queryset.count()}")
        print(f"Page object count: {len(page_obj)}")
        
        context = {
            'page_obj': page_obj,  # Use page_obj for pagination
            'search_query': search_query,
            'gender_filter': gender_filter,
            'dusun_filter': dusun_filter,
            'marital_status_filter': marital_status_filter,
            'dusun_options': dusun_options,
            'total_penduduk': penduduk_queryset.count(),
            'debug_info': {
                'total_count': total_count,
                'filtered_count': penduduk_queryset.count(),
                'page_count': len(page_obj),
            }
        }
        return render(request, 'admin_panel/references/penduduk_list.html', context)
    except Exception as e:
        print(f"Error in penduduk_list: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error loading penduduk data: {str(e)}')
        return redirect('admin_panel:references_dashboard')

@login_required
def penduduk_detail(request, penduduk_id):
    """View untuk menampilkan detail penduduk"""
    try:
        penduduk = get_object_or_404(Penduduk, id=penduduk_id)
        context = {
            'penduduk': penduduk
        }
        return render(request, 'admin_panel/references/detail/penduduk_detail.html', context)
    except Exception as e:
        messages.error(request, f'Error loading penduduk detail: {str(e)}')
        return redirect('admin_panel:references_penduduk_list')

@login_required
def penduduk_add(request):
    """View untuk menambah data penduduk baru"""
    try:
        dusun_list = Dusun.objects.filter(is_active=True)
        context = {
            'dusun_list': dusun_list,
            'relationship_choices': Penduduk.RELATIONSHIP_CHOICES,
            'gender_choices': Penduduk.GENDER_CHOICES,
            'religion_choices': Penduduk.RELIGION_CHOICES,
            'marital_status_choices': Penduduk.MARITAL_STATUS_CHOICES,
            'education_choices': Penduduk.EDUCATION_CHOICES,
            'blood_type_choices': Penduduk.BLOOD_TYPE_CHOICES,
            'citizenship_choices': Penduduk.CITIZENSHIP_CHOICES,
            'form_title': 'Tambah Penduduk',
            'form_subtitle': 'Lengkapi data penduduk baru',
            'form_icon': 'plus'
        }
        return render(request, 'admin_panel/references/penduduk_form.html', context)
    except Exception as e:
        messages.error(request, f'Error loading form: {str(e)}')
        return redirect('admin_panel:references_penduduk_list')

@login_required
def penduduk_create(request):
    """View untuk membuat data penduduk baru"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            penduduk = Penduduk.objects.create(
                name=data['name'],
                nik=data['nik'],
                dusun_id=data['dusun_id'],
                keluarga_id=data.get('keluarga_id'),
                is_active=True,
                created_by=request.user
            )
            return JsonResponse({
                'success': True,
                'data': {
                    'id': penduduk.id,
                    'name': penduduk.name,
                    'nik': penduduk.nik
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def penduduk_detail(request, penduduk_id):
    """View untuk menampilkan detail penduduk"""
    try:
        penduduk = get_object_or_404(Penduduk, id=penduduk_id)
        context = {
            'penduduk': penduduk
        }
        return render(request, 'admin_panel/references/penduduk_detail.html', context)
    except Penduduk.DoesNotExist:
        messages.error(request, 'Data penduduk tidak ditemukan.')
        return redirect('admin_panel:references_penduduk_list')

@login_required
def penduduk_edit(request, penduduk_id):
    """View untuk mengedit penduduk"""
    try:
        penduduk = get_object_or_404(Penduduk, id=penduduk_id)
        dusun_list = Dusun.objects.filter(is_active=True)
        
        # Get RT and RW data if they exist
        rt_data = None
        rw_data = None
        if penduduk.rt:
            rt_data = {
                'id': penduduk.rt.id,
                'rt_number': penduduk.rt.rt_number
            }
        if penduduk.rw:
            rw_data = {
                'id': penduduk.rw.id,
                'rw_number': penduduk.rw.rw_number
            }
        
        context = {
            'penduduk': penduduk,
            'dusun_list': dusun_list,
            'relationship_choices': Penduduk.RELATIONSHIP_CHOICES,
            'gender_choices': Penduduk.GENDER_CHOICES,
            'religion_choices': Penduduk.RELIGION_CHOICES,
            'marital_status_choices': Penduduk.MARITAL_STATUS_CHOICES,
            'education_choices': Penduduk.EDUCATION_CHOICES,
            'blood_type_choices': Penduduk.BLOOD_TYPE_CHOICES,
            'citizenship_choices': Penduduk.CITIZENSHIP_CHOICES,
            'rt_data': rt_data,
            'rw_data': rw_data,
            'form_title': 'Edit Penduduk',
            'form_subtitle': 'Ubah data penduduk',
            'form_icon': 'edit'
        }
        return render(request, 'admin_panel/references/penduduk_form.html', context)
    except Penduduk.DoesNotExist:
        messages.error(request, 'Data penduduk tidak ditemukan.')
        return redirect('admin_panel:references_penduduk_list')

@login_required
def penduduk_update(request, penduduk_id):
    """View untuk mengupdate data penduduk"""
    try:
        penduduk = get_object_or_404(Penduduk, id=penduduk_id)
        if request.method == 'POST':
            data = json.loads(request.body)
            penduduk.name = data['name']
            penduduk.nik = data['nik']
            penduduk.dusun_id = data['dusun_id']
            penduduk.keluarga_id = data.get('keluarga_id')
            penduduk.updated_by = request.user
            penduduk.save()
            return JsonResponse({'success': True})
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    except Penduduk.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Penduduk not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def penduduk_delete(request, penduduk_id):
    """View untuk menghapus penduduk via modal"""
    if request.method == 'POST':
        try:
            penduduk = get_object_or_404(Penduduk, id=penduduk_id)
            penduduk_name = penduduk.name
            
            # Get reason from form data
            reason = request.POST.get('reason', '')
            
            # Delete the penduduk
            penduduk.delete()
            
            # Return JSON response for modal
            return JsonResponse({
                'success': True, 
                'message': f'Data penduduk {penduduk_name} berhasil dihapus'
            })
            
        except Penduduk.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Data penduduk tidak ditemukan'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    else:
        # For GET requests, redirect to list
        return redirect('admin_panel:references_penduduk_list')

@login_required
def penduduk_save(request):
    """View untuk menyimpan data penduduk"""
    if request.method == 'POST':
        try:
            # Get form data
            nik = request.POST.get('nik')
            name = request.POST.get('nama_lengkap')
            birth_place = request.POST.get('tempat_lahir')
            birth_date = request.POST.get('tanggal_lahir')
            gender = request.POST.get('jenis_kelamin')
            religion = request.POST.get('agama')
            marital_status = request.POST.get('status_perkawinan')
            education = request.POST.get('education')
            occupation = request.POST.get('pekerjaan')
            address = request.POST.get('alamat')
            dusun_id = request.POST.get('dusun')
            lorong_id = request.POST.get('lorong')
            rt_number = request.POST.get('rt_number')
            rw_number = request.POST.get('rw_number')
            house_number = request.POST.get('house_number')
            postal_code = request.POST.get('postal_code')
            phone_number = request.POST.get('phone_number')
            mobile_number = request.POST.get('mobile_number')
            email = request.POST.get('email')
            blood_type = request.POST.get('blood_type')
            kk_number = request.POST.get('kk_number')
            relationship_to_head = request.POST.get('relationship_to_head')
            is_active = request.POST.get('is_active') == 'on'
            
            # Handle photo upload
            photo = request.FILES.get('photo')
            
            # Validate photo if provided
            if photo:
                # Check file size (max 2MB)
                if photo.size > 2 * 1024 * 1024:
                    messages.error(request, 'Ukuran foto maksimal 2MB.')
                    return redirect(request.META.get('HTTP_REFERER', 'admin_panel:references_penduduk_list'))
                
                # Check file type
                allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
                if photo.content_type not in allowed_types:
                    messages.error(request, 'Format foto harus JPG, PNG, atau WebP.')
                    return redirect(request.META.get('HTTP_REFERER', 'admin_panel:references_penduduk_list'))
            
            # Validate required fields
            required_fields = {
                'nik': nik,
                'nama_lengkap': name,
                'tempat_lahir': birth_place,
                'tanggal_lahir': birth_date,
                'jenis_kelamin': gender,
                'agama': religion,
                'status_perkawinan': marital_status,
                'alamat': address,
                'dusun': dusun_id
            }
            
            # Check if KK is required for kepala keluarga
            if relationship_to_head == 'KEPALA_KELUARGA':
                required_fields['kk_number'] = kk_number
            
            missing_fields = [field for field, value in required_fields.items() if not value]
            if missing_fields:
                messages.error(request, f'Field yang wajib diisi: {", ".join(missing_fields)}')
                return redirect(request.META.get('HTTP_REFERER', 'admin_panel:references_penduduk_list'))
            
            # Validate NIK format and uniqueness
            if nik and (not nik.isdigit() or len(nik) != 16):
                messages.error(request, 'NIK harus berupa 16 digit angka.')
                return redirect(request.META.get('HTTP_REFERER', 'admin_panel:references_penduduk_list'))
            
            # Get penduduk_id for edit mode
            penduduk_id = request.POST.get('penduduk_id')
            
            # Check NIK uniqueness
            if penduduk_id:
                # Update mode - check if NIK exists for other records
                existing_penduduk = Penduduk.objects.filter(nik=nik).exclude(id=penduduk_id).first()
            else:
                # Create mode - check if NIK exists
                existing_penduduk = Penduduk.objects.filter(nik=nik).first()
            
            if existing_penduduk:
                messages.error(request, f'NIK {nik} sudah terdaftar untuk {existing_penduduk.name}.')
                return redirect(request.META.get('HTTP_REFERER', 'admin_panel:references_penduduk_list'))
            
            if penduduk_id:
                # Update existing
                penduduk = get_object_or_404(Penduduk, id=penduduk_id)
                penduduk.nik = nik
                penduduk.name = name
                penduduk.birth_place = birth_place
                penduduk.birth_date = birth_date
                penduduk.gender = gender
                penduduk.religion = religion
                penduduk.marital_status = marital_status
                penduduk.education = education
                penduduk.occupation = occupation
                penduduk.address = address
                penduduk.dusun_id = dusun_id
                penduduk.lorong_id = lorong_id if lorong_id else None
                penduduk.rt_number = rt_number
                penduduk.rw_number = rw_number
                penduduk.house_number = house_number
                penduduk.postal_code = postal_code
                penduduk.phone_number = phone_number
                penduduk.mobile_number = mobile_number
                penduduk.email = email
                penduduk.blood_type = blood_type
                penduduk.kk_number = kk_number
                penduduk.relationship_to_head = relationship_to_head
                penduduk.is_active = is_active
                penduduk.updated_by = request.user
                
                # Update photo if provided
                if photo:
                    penduduk.photo = photo
                
                penduduk.save()
                
                # Auto-create keluarga if kepala keluarga
                if relationship_to_head == 'KEPALA_KELUARGA' and kk_number:
                    # TODO: Implement keluarga creation logic
                    pass
                
                messages.success(request, f'Data penduduk {name} berhasil diperbarui.')
            else:
                # Create new
                penduduk = Penduduk.objects.create(
                    nik=nik,
                    name=name,
                    birth_place=birth_place,
                    birth_date=birth_date,
                    gender=gender,
                    religion=religion,
                    marital_status=marital_status,
                    education=education,
                    occupation=occupation,
                    address=address,
                    dusun_id=dusun_id,
                    lorong_id=lorong_id if lorong_id else None,
                    rt_number=rt_number,
                    rw_number=rw_number,
                    house_number=house_number,
                    postal_code=postal_code,
                    phone_number=phone_number,
                    mobile_number=mobile_number,
                    email=email,
                    blood_type=blood_type,
                    kk_number=kk_number,
                    relationship_to_head=relationship_to_head,
                    is_active=is_active,
                    photo=photo if photo else None,
                    created_by=request.user
                )
                
                # Auto-create keluarga if kepala keluarga
                if relationship_to_head == 'KEPALA_KELUARGA' and kk_number:
                    # TODO: Implement keluarga creation logic
                    pass
                
                messages.success(request, f'Data penduduk {name} berhasil ditambahkan.')
            
            return redirect('admin_panel:references_penduduk_list')
            
        except Exception as e:
            messages.error(request, f'Error menyimpan data: {str(e)}')
            return redirect('admin_panel:references_penduduk_list')
    
    return redirect('admin_panel:references_penduduk_list')

@login_required
def penduduk_upload_photo(request):
    """View untuk upload foto penduduk via AJAX"""
    if request.method == 'POST':
        try:
            penduduk_id = request.POST.get('penduduk_id')
            photo = request.FILES.get('photo')
            
            if not penduduk_id or not photo:
                return JsonResponse({
                    'success': False,
                    'message': 'Penduduk ID dan foto harus diisi'
                })
            
            # Validate file size (max 2MB)
            if photo.size > 2 * 1024 * 1024:
                return JsonResponse({
                    'success': False,
                    'message': 'Ukuran foto maksimal 2MB'
                })
            
            # Validate file type
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
            if photo.content_type not in allowed_types:
                return JsonResponse({
                    'success': False,
                    'message': 'Format foto harus JPG, PNG, atau WebP'
                })
            
            # Get penduduk
            penduduk = get_object_or_404(Penduduk, id=penduduk_id)
            
            # Update photo
            penduduk.photo = photo
            penduduk.updated_by = request.user
            penduduk.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Foto berhasil diupload',
                'photo_url': penduduk.photo.url
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Method not allowed'
    })

# Import/Export functionality moved to views_import_export.py

# RW Management Functions
@login_required
def rw_list(request):
    """List all RW data"""
    rws = RW.objects.select_related('dusun', 'ketua_rw').all().order_by('dusun__name', 'rw_number')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        rws = rws.filter(
            Q(rw_number__icontains=search_query) |
            Q(dusun__name__icontains=search_query) |
            Q(ketua_rw__name__icontains=search_query)
        )
    
    # Filter by dusun
    dusun_filter = request.GET.get('dusun', '')
    if dusun_filter:
        rws = rws.filter(dusun_id=dusun_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        rws = rws.filter(is_active=True)
    elif status_filter == 'inactive':
        rws = rws.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(rws, 20)
    page_number = request.GET.get('page')
    rws = paginator.get_page(page_number)
    
    # Get dusun list for filter
    dusuns = Dusun.objects.all()
    
    context = {
        'title': 'Data RW',
        'page_obj': rws,
        'dusuns': dusuns,
        'search_query': search_query,
        'dusun_filter': dusun_filter,
        'status_filter': status_filter,
        'total_count': RW.objects.count(),
    }
    return render(request, 'admin_panel/references/rw_list.html', context)

@login_required
def rw_create(request):
    """Create new RW"""
    if request.method == 'POST':
        form = RWForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'RW berhasil dibuat')
            return redirect('admin_panel:references_rw_list')
    else:
        form = RWForm()
    
    context = {
        'title': 'Tambah RW',
        'form': form,
    }
    return render(request, 'admin_panel/references/rw_form.html', context)

@login_required
def rw_edit(request, rw_id):
    """Edit RW"""
    try:
        rw = RW.objects.get(id=rw_id)
    except RW.DoesNotExist:
        messages.error(request, 'RW tidak ditemukan')
        return redirect('admin_panel:references_rw_list')
    
    if request.method == 'POST':
        form = RWForm(request.POST, instance=rw)
        if form.is_valid():
            form.save()
            messages.success(request, 'RW berhasil diperbarui')
            return redirect('admin_panel:references_rw_list')
    else:
        form = RWForm(instance=rw)
    
    context = {
        'title': 'Edit RW',
        'form': form,
        'rw': rw,
    }
    return render(request, 'admin_panel/references/rw_form.html', context)

@login_required
def rw_delete(request, rw_id):
    """Delete RW"""
    try:
        rw = RW.objects.get(id=rw_id)
        rw.delete()
        messages.success(request, 'RW berhasil dihapus')
    except RW.DoesNotExist:
        messages.error(request, 'RW tidak ditemukan')
    
    return redirect('admin_panel:references_rw_list')

# RT Management Functions
@login_required
def rt_list(request):
    """List all RT data"""
    rts = RT.objects.select_related('rw__dusun', 'ketua_rt').all().order_by('rw__dusun__name', 'rw__rw_number', 'rt_number')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        rts = rts.filter(
            Q(rt_number__icontains=search_query) |
            Q(rw__rw_number__icontains=search_query) |
            Q(rw__dusun__name__icontains=search_query) |
            Q(ketua_rt__name__icontains=search_query)
        )
    
    # Filter by RW
    rw_filter = request.GET.get('rw', '')
    if rw_filter:
        rts = rts.filter(rw_id=rw_filter)
    
    # Filter by dusun
    dusun_filter = request.GET.get('dusun', '')
    if dusun_filter:
        rts = rts.filter(rw__dusun_id=dusun_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        rts = rts.filter(is_active=True)
    elif status_filter == 'inactive':
        rts = rts.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(rts, 20)
    page_number = request.GET.get('page')
    rts = paginator.get_page(page_number)
    
    # Get RW and dusun lists for filter
    rws = RW.objects.select_related('dusun').all()
    dusuns = Dusun.objects.all()
    
    context = {
        'title': 'Data RT',
        'page_obj': rts,
        'rws': rws,
        'dusuns': dusuns,
        'search_query': search_query,
        'rw_filter': rw_filter,
        'dusun_filter': dusun_filter,
        'status_filter': status_filter,
        'total_count': RT.objects.count(),
    }
    return render(request, 'admin_panel/references/rt_list.html', context)

@login_required
def rt_create(request):
    """Create new RT"""
    if request.method == 'POST':
        form = RTForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'RT berhasil dibuat')
            return redirect('admin_panel:references_rt_list')
    else:
        form = RTForm()
    
    context = {
        'title': 'Tambah RT',
        'form': form,
    }
    return render(request, 'admin_panel/references/rt_form.html', context)

@login_required
def rt_edit(request, rt_id):
    """Edit RT"""
    try:
        rt = RT.objects.get(id=rt_id)
    except RT.DoesNotExist:
        messages.error(request, 'RT tidak ditemukan')
        return redirect('admin_panel:references_rt_list')
    
    if request.method == 'POST':
        form = RTForm(request.POST, instance=rt)
        if form.is_valid():
            form.save()
            messages.success(request, 'RT berhasil diperbarui')
            return redirect('admin_panel:references_rt_list')
    else:
        form = RTForm(instance=rt)
    
    context = {
        'title': 'Edit RT',
        'form': form,
        'rt': rt,
    }
    return render(request, 'admin_panel/references/rt_form.html', context)

@login_required
def rt_delete(request, rt_id):
    """Delete RT"""
    try:
        rt = RT.objects.get(id=rt_id)
        rt.delete()
        messages.success(request, 'RT berhasil dihapus')
    except RT.DoesNotExist:
        messages.error(request, 'RT tidak ditemukan')
    
    return redirect('admin_panel:references_rt_list')

# API Functions for RW and RT
@login_required
def api_rw_by_dusun(request, dusun_id):
    """Get RW list by dusun ID"""
    try:
        rws = RW.objects.filter(dusun_id=dusun_id, is_active=True).select_related('dusun', 'ketua_rw')
        data = []
        for rw in rws:
            data.append({
                'id': rw.id,
                'rw_number': rw.rw_number,
                'dusun_name': rw.dusun.name,
                'ketua_rw': rw.ketua_rw.name if rw.ketua_rw else None,
                'population_count': rw.population_count,
            })
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_rt_by_rw(request, rw_id):
    """Get RT list by RW ID"""
    try:
        rts = RT.objects.filter(rw_id=rw_id, is_active=True).select_related('rw__dusun', 'ketua_rt')
        data = []
        for rt in rts:
            data.append({
                'id': rt.id,
                'rt_number': rt.rt_number,
                'rw_number': rt.rw.rw_number,
                'dusun_name': rt.rw.dusun.name,
                'ketua_rt': rt.ketua_rt.name if rt.ketua_rt else None,
                'population_count': rt.population_count,
            })
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_penduduk_by_rw(request, rw_id):
    """Get penduduk list by RW ID"""
    try:
        penduduks = Penduduk.objects.filter(
            rw_number__in=RT.objects.filter(rw_id=rw_id).values_list('rt_number', flat=True),
            is_active=True
        ).select_related('dusun', 'lorong')
        data = []
        for penduduk in penduduks:
            data.append({
                'id': penduduk.id,
                'name': penduduk.name,
                'nik': penduduk.nik,
                'gender': penduduk.get_gender_display(),
                'birth_date': penduduk.birth_date.strftime('%Y-%m-%d') if penduduk.birth_date else None,
                'dusun': penduduk.dusun.name if penduduk.dusun else None,
                'rt_number': penduduk.rt_number,
                'rw_number': penduduk.rw_number,
            })
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_penduduk_by_dusun(request, dusun_id):
    """Get penduduk list by dusun ID"""
    try:
        penduduks = Penduduk.objects.filter(
            dusun_id=dusun_id,
                        is_active=True
        ).select_related('dusun', 'lorong')
        data = []
        for penduduk in penduduks:
            data.append({
                'id': penduduk.id,
                'name': penduduk.name,
                'nik': penduduk.nik,
                'gender': penduduk.get_gender_display(),
                'birth_date': penduduk.birth_date.strftime('%Y-%m-%d') if penduduk.birth_date else None,
                'dusun': penduduk.dusun.name if penduduk.dusun else None,
                'rt_number': penduduk.rt_number,
                'rw_number': penduduk.rw_number,
            })
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_penduduk_list(request):
    """Get all penduduk list"""
    try:
        penduduks = Penduduk.objects.filter(is_active=True).select_related('dusun', 'lorong')                                                                   
        data = []
        for penduduk in penduduks:
            data.append({
                'id': penduduk.id,
                'name': penduduk.name,
                'nik': penduduk.nik,
                'gender': penduduk.get_gender_display(),
                'birth_date': penduduk.birth_date.strftime('%Y-%m-%d') if penduduk.birth_date else None,                                                        
                'dusun': penduduk.dusun.name if penduduk.dusun else None,       
                'rt_number': penduduk.rt_number or '',
                'rw_number': penduduk.rw_number or '',
            })
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# RT Bulk Actions
@login_required
def rt_bulk_delete(request):
    """Bulk delete RT records"""
    if request.method == 'POST':
        try:
            rt_ids = request.POST.getlist('rt_ids[]')
            if not rt_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            rt_ids = [int(rt_id) for rt_id in rt_ids]
            
            # Get RT objects
            rt_objects = RT.objects.filter(id__in=rt_ids)
            deleted_count = rt_objects.count()
            
            # Delete the objects
            rt_objects.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menghapus {deleted_count} data RT',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def rt_bulk_activate(request):
    """Bulk activate RT records"""
    if request.method == 'POST':
        try:
            rt_ids = request.POST.getlist('rt_ids[]')
            if not rt_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            rt_ids = [int(rt_id) for rt_id in rt_ids]
            
            # Update RT objects
            updated_count = RT.objects.filter(id__in=rt_ids).update(is_active=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil mengaktifkan {updated_count} data RT',
                'activated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def rt_bulk_deactivate(request):
    """Bulk deactivate RT records"""
    if request.method == 'POST':
        try:
            rt_ids = request.POST.getlist('rt_ids[]')
            if not rt_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            rt_ids = [int(rt_id) for rt_id in rt_ids]
            
            # Update RT objects
            updated_count = RT.objects.filter(id__in=rt_ids).update(is_active=False)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menonaktifkan {updated_count} data RT',
                'deactivated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


# Keluarga Bulk Actions
@login_required
def keluarga_bulk_delete(request):
    """Bulk delete Keluarga records"""
    if request.method == 'POST':
        try:
            keluarga_ids = request.POST.getlist('keluarga_ids[]')
            if not keluarga_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            keluarga_ids = [int(keluarga_id) for keluarga_id in keluarga_ids]
            
            # Get Keluarga objects
            keluarga_objects = Keluarga.objects.filter(id__in=keluarga_ids)
            deleted_count = keluarga_objects.count()
            
            # Delete the objects
            keluarga_objects.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menghapus {deleted_count} data Keluarga',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def keluarga_bulk_activate(request):
    """Bulk activate Keluarga records"""
    if request.method == 'POST':
        try:
            keluarga_ids = request.POST.getlist('keluarga_ids[]')
            if not keluarga_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            keluarga_ids = [int(keluarga_id) for keluarga_id in keluarga_ids]
            
            # Update Keluarga objects
            updated_count = Keluarga.objects.filter(id__in=keluarga_ids).update(is_active=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil mengaktifkan {updated_count} data Keluarga',
                'activated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def keluarga_bulk_deactivate(request):
    """Bulk deactivate Keluarga records"""
    if request.method == 'POST':
        try:
            keluarga_ids = request.POST.getlist('keluarga_ids[]')
            if not keluarga_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            keluarga_ids = [int(keluarga_id) for keluarga_id in keluarga_ids]
            
            # Update Keluarga objects
            updated_count = Keluarga.objects.filter(id__in=keluarga_ids).update(is_active=False)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menonaktifkan {updated_count} data Keluarga',
                'deactivated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


# Template Download Functions


# RW Bulk Actions
@login_required
def rw_bulk_delete(request):
    """Bulk delete RW records"""
    if request.method == 'POST':
        try:
            rw_ids = request.POST.getlist('rw_ids[]')
            if not rw_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            rw_ids = [int(rw_id) for rw_id in rw_ids]
            
            # Get RW objects
            rw_objects = RW.objects.filter(id__in=rw_ids)
            deleted_count = rw_objects.count()
            
            # Delete the objects
            rw_objects.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menghapus {deleted_count} data RW',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def rw_bulk_activate(request):
    """Bulk activate RW records"""
    if request.method == 'POST':
        try:
            rw_ids = request.POST.getlist('rw_ids[]')
            if not rw_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            rw_ids = [int(rw_id) for rw_id in rw_ids]
            
            # Update RW objects
            updated_count = RW.objects.filter(id__in=rw_ids).update(is_active=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil mengaktifkan {updated_count} data RW',
                'activated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def rw_bulk_deactivate(request):
    """Bulk deactivate RW records"""
    if request.method == 'POST':
        try:
            rw_ids = request.POST.getlist('rw_ids[]')
            if not rw_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            # Convert to integers
            rw_ids = [int(rw_id) for rw_id in rw_ids]
            
            # Update RW objects
            updated_count = RW.objects.filter(id__in=rw_ids).update(is_active=False)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menonaktifkan {updated_count} data RW',
                'deactivated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

def api_keluarga_list(request):
    """API untuk mendapatkan daftar keluarga"""
    try:
        from django.core.paginator import Paginator
        
        keluarga_list = Keluarga.objects.select_related('dusun').filter(is_active=True).order_by('nama_kepala_keluarga')
        
        search = request.GET.get('search')
        dusun = request.GET.get('dusun')
        rt = request.GET.get('rt')
        rw = request.GET.get('rw')
        
        if search:
            keluarga_list = keluarga_list.filter(
                Q(nama_kepala_keluarga__icontains=search) |
                Q(nomor_kk__icontains=search) |
                Q(alamat__icontains=search)
            )
        
        if dusun:
            keluarga_list = keluarga_list.filter(dusun_id=dusun)
            
        if rt:
            keluarga_list = keluarga_list.filter(rt=rt)
            
        if rw:
            keluarga_list = keluarga_list.filter(rw=rw)
        
        page = int(request.GET.get('page', 1))
        per_page = 15
        paginator = Paginator(keluarga_list, per_page)
        page_obj = paginator.get_page(page)
        
        total_keluarga = keluarga_list.count()
        active_keluarga = keluarga_list.filter(is_active=True).count()
        
        results = [{
            'id': keluarga.id,
            'nomor_kk': keluarga.nomor_kk,
            'nama_kepala_keluarga': keluarga.nama_kepala_keluarga,
            'alamat': keluarga.alamat,
            'dusun_id': keluarga.dusun.id if keluarga.dusun else None,
            'dusun_name': keluarga.dusun.name if keluarga.dusun else '',
            'rt': keluarga.rt or '',
            'rw': keluarga.rw or '',
            'is_active': keluarga.is_active,
            'created_at': keluarga.created_at.isoformat() if keluarga.created_at else None,
            'updated_at': keluarga.updated_at.isoformat() if keluarga.updated_at else None
        } for keluarga in page_obj]
        
        return JsonResponse({
            'success': True,
            'data': results,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'per_page': per_page,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None
            },
            'statistics': {
                'total_keluarga': total_keluarga,
                'active_keluarga': active_keluarga
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# Pelajar Bulk Actions
@login_required
def pelajar_bulk_delete(request):
    """Bulk delete Pelajar records"""
    if request.method == 'POST':
        try:
            pelajar_ids = request.POST.getlist('pelajar_ids[]')
            if not pelajar_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            pelajar_ids = [int(pelajar_id) for pelajar_id in pelajar_ids]
            pelajar_objects = Pelajar.objects.filter(id__in=pelajar_ids)
            deleted_count = pelajar_objects.count()
            pelajar_objects.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menghapus {deleted_count} data Pelajar',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def pelajar_bulk_activate(request):
    """Bulk activate Pelajar records"""
    if request.method == 'POST':
        try:
            pelajar_ids = request.POST.getlist('pelajar_ids[]')
            if not pelajar_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            pelajar_ids = [int(pelajar_id) for pelajar_id in pelajar_ids]
            updated_count = Pelajar.objects.filter(id__in=pelajar_ids).update(is_active=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil mengaktifkan {updated_count} data Pelajar',
                'activated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def pelajar_bulk_deactivate(request):
    """Bulk deactivate Pelajar records"""
    if request.method == 'POST':
        try:
            pelajar_ids = request.POST.getlist('pelajar_ids[]')
            if not pelajar_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            pelajar_ids = [int(pelajar_id) for pelajar_id in pelajar_ids]
            updated_count = Pelajar.objects.filter(id__in=pelajar_ids).update(is_active=False)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menonaktifkan {updated_count} data Pelajar',
                'deactivated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

# Penduduk Bulk Actions
@login_required
def penduduk_bulk_delete(request):
    """Bulk delete Penduduk records"""
    if request.method == 'POST':
        try:
            penduduk_ids = request.POST.getlist('penduduk_ids[]')
            if not penduduk_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            penduduk_ids = [int(penduduk_id) for penduduk_id in penduduk_ids]
            penduduk_objects = Penduduk.objects.filter(id__in=penduduk_ids)
            deleted_count = penduduk_objects.count()
            penduduk_objects.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menghapus {deleted_count} data Penduduk',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def penduduk_bulk_activate(request):
    """Bulk activate Penduduk records"""
    if request.method == 'POST':
        try:
            penduduk_ids = request.POST.getlist('penduduk_ids[]')
            if not penduduk_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            penduduk_ids = [int(penduduk_id) for penduduk_id in penduduk_ids]
            updated_count = Penduduk.objects.filter(id__in=penduduk_ids).update(is_active=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil mengaktifkan {updated_count} data Penduduk',
                'activated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def penduduk_bulk_deactivate(request):
    """Bulk deactivate Penduduk records"""
    if request.method == 'POST':
        try:
            penduduk_ids = request.POST.getlist('penduduk_ids[]')
            if not penduduk_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            penduduk_ids = [int(penduduk_id) for penduduk_id in penduduk_ids]
            updated_count = Penduduk.objects.filter(id__in=penduduk_ids).update(is_active=False)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menonaktifkan {updated_count} data Penduduk',
                'deactivated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

# Disabilitas Bulk Actions
@login_required
def disabilitas_bulk_delete(request):
    """Bulk delete DisabilitasData records"""
    if request.method == 'POST':
        try:
            disabilitas_ids = request.POST.getlist('disabilitas_ids[]')
            if not disabilitas_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            disabilitas_ids = [int(disabilitas_id) for disabilitas_id in disabilitas_ids]
            disabilitas_objects = DisabilitasData.objects.filter(id__in=disabilitas_ids)
            deleted_count = disabilitas_objects.count()
            disabilitas_objects.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menghapus {deleted_count} data Disabilitas',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def disabilitas_bulk_activate(request):
    """Bulk activate DisabilitasData records"""
    if request.method == 'POST':
        try:
            disabilitas_ids = request.POST.getlist('disabilitas_ids[]')
            if not disabilitas_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            disabilitas_ids = [int(disabilitas_id) for disabilitas_id in disabilitas_ids]
            updated_count = DisabilitasData.objects.filter(id__in=disabilitas_ids).update(is_active=True)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil mengaktifkan {updated_count} data Disabilitas',
                'activated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def disabilitas_bulk_deactivate(request):
    """Bulk deactivate DisabilitasData records"""
    if request.method == 'POST':
        try:
            disabilitas_ids = request.POST.getlist('disabilitas_ids[]')
            if not disabilitas_ids:
                return JsonResponse({'success': False, 'message': 'Tidak ada data yang dipilih'})
            
            disabilitas_ids = [int(disabilitas_id) for disabilitas_id in disabilitas_ids]
            updated_count = DisabilitasData.objects.filter(id__in=disabilitas_ids).update(is_active=False)
            
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil menonaktifkan {updated_count} data Disabilitas',
                'deactivated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

# API RT List
def api_rt_list(request):
    """API untuk mendapatkan daftar RT"""
    try:
        from django.core.paginator import Paginator
        
        rt_list = RT.objects.select_related('rw').filter(is_active=True).order_by('rt_number')
        
        search = request.GET.get('search')
        rw = request.GET.get('rw')
        
        if search:
            rt_list = rt_list.filter(
                Q(rt_number__icontains=search) |
                Q(ketua_rt__name__icontains=search) |
                Q(alamat__icontains=search)
            )
        
        if rw:
            rt_list = rt_list.filter(rw_id=rw)
        
        page = int(request.GET.get('page', 1))
        per_page = 15
        paginator = Paginator(rt_list, per_page)
        page_obj = paginator.get_page(page)
        
        total_rt = rt_list.count()
        active_rt = rt_list.filter(is_active=True).count()
        
        results = [{
            'id': rt.id,
            'rt_number': rt.rt_number,
            'ketua_rt': rt.ketua_rt.name if rt.ketua_rt else '',
            'alamat': rt.alamat,
            'rw_id': rt.rw.id if rt.rw else None,
            'rw_name': rt.rw.rw_number if rt.rw else '',
            'is_active': rt.is_active,
            'created_at': rt.created_at.isoformat() if rt.created_at else None,
            'updated_at': rt.updated_at.isoformat() if rt.updated_at else None
        } for rt in page_obj]
        
        return JsonResponse({
            'success': True,
            'data': results,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'per_page': per_page,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None
            },
            'statistics': {
                'total_rt': total_rt,
                'active_rt': active_rt
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# API RW List
def api_rw_list(request):
    """API untuk mendapatkan daftar RW"""
    try:
        from django.core.paginator import Paginator
        
        rw_list = RW.objects.filter(is_active=True).order_by('rw_number')
        
        search = request.GET.get('search')
        
        if search:
            rw_list = rw_list.filter(
                Q(rw_number__icontains=search) |
                Q(ketua_rw__name__icontains=search) |
                Q(alamat__icontains=search)
            )
        
        page = int(request.GET.get('page', 1))
        per_page = 15
        paginator = Paginator(rw_list, per_page)
        page_obj = paginator.get_page(page)
        
        total_rw = rw_list.count()
        active_rw = rw_list.filter(is_active=True).count()
        
        results = [{
            'id': rw.id,
            'rw_number': rw.rw_number,
            'ketua_rw': rw.ketua_rw.name if rw.ketua_rw else '',
            'alamat': rw.alamat,
            'is_active': rw.is_active,
            'created_at': rw.created_at.isoformat() if rw.created_at else None,
            'updated_at': rw.updated_at.isoformat() if rw.updated_at else None
        } for rw in page_obj]
        
        return JsonResponse({
            'success': True,
            'data': results,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'per_page': per_page,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None
            },
            'statistics': {
                'total_rw': total_rw,
                'active_rw': active_rw
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
